import openpyxl

def excel_ausgabe(daten_roh, verzeichnis):
	
	wb = workbook_erstellen()
	wb = daten_sheet(wb, daten_roh)
	workbook_speichern(wb, verzeichnis)


def workbook_erstellen():	
	wb = openpyxl.Workbook()

	return wb


def daten_sheet(wb, daten_roh):

	if len(daten_roh) == 12:
		monate = ['Januar','Februar','März','April','Mai','Juni','Juli','August','September','Oktober','November','Dezember']
	else:
		monate = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag']

	for i in range(len(daten_roh)):
		if i ==0:
			ws = wb.active
			ws.titel = monate[i]
		else:
			wb.create_sheet(monate[i])
			ws = wb[monate[i]]

		da = daten_roh[i][5]
		op = daten_roh[i][1]
		low = daten_roh[i][6]
		high = daten_roh[i][7]
		clo = daten_roh[i][2]
		dif = daten_roh[i][3]
		span = daten_roh[i][4]

		ws.cell(row = 1, column = 1, value = 'Date')
		ws.cell(row = 1, column = 2, value = 'Open')
		ws.cell(row = 1, column = 3, value = 'Low')
		ws.cell(row = 1, column = 4, value = 'High')
		ws.cell(row = 1, column = 5, value = 'Close')
		ws.cell(row = 1, column = 7, value = 'stet rend')
		ws.cell(row = 1, column = 8, value = 'Spanne')

		for i in range(len(daten_roh[0])):
			ws.cell(row = i+2, column = 1, value = da[i])
			ws.cell(row = i+2, column = 2, value = op[i])
			ws.cell(row = i+2, column = 3, value = low[i])
			ws.cell(row = i+2, column = 4, value = high[i])
			ws.cell(row = i+2, column = 5, value = clo[i])
			ws.cell(row = i+2, column = 7, value = dif[i])
			ws.cell(row = i+2, column = 8, value = span[i])

	return wb

def workbook_speichern(w, verzeichnis):
	wb = w
	
	if not os.path.exists(verzeichnis):
		os.makedirs(verzeichnis)

	wb.save(verzeichnis +r'\_time-value_Überprüfung.xlsx')