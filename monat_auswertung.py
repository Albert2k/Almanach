import openpyxl
import numpy as np
from scipy import stats as scistats
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import collections
import matplotlib.mlab as mlab
import os
import math

def daten_einlesen(excel_datei):
	ws = excel_einlesen(excel_datei)
	rohdaten = rohdaten_einlesen_verarbeiten(ws)
	print('Daten eingelesen!')

	return rohdaten


def excel_einlesen(excel_datei):

	wb_daten = openpyxl.load_workbook(filename = excel_datei)
	ws_daten = wb_daten.active

	return [ws_daten]

#--------------------------------------------------------------------------------------------------------------------------
# Retrun-Liste: [0] Datum; [1] Open; [2] High; [3] Low; [4] Close; [5] Differenz; [6] höchste Schwankung

def rohdaten_einlesen_verarbeiten(ws_rohdaten):

	daten_datum = spalte_einlesen(ws_rohdaten, 'A', 'Date')
	daten_open = spalte_einlesen(ws_rohdaten, 'B', 'Open')
	daten_high = spalte_einlesen(ws_rohdaten, 'C', 'High')
	daten_low = spalte_einlesen(ws_rohdaten, 'D', 'Low')
	daten_close = spalte_einlesen(ws_rohdaten, 'E', 'Close')

	daten_diff = []
	for i in range(len(daten_open)):
		daten_diff = daten_diff + [round((daten_close[i] - daten_open[i])/daten_open[i]*100,2)]

	daten_schw = []
	for i in range(len(daten_high)):
		daten_schw = daten_schw + [round((daten_high[i] - daten_low[i])/daten_open[i]*100,2)]

	return [daten_datum, daten_open, daten_high, daten_low, daten_close, daten_diff, daten_schw]



def spalte_einlesen(ws_rohdaten, spalte_name, name):
	ws = ws_rohdaten[0]
	spalten = ws[spalte_name]
	spalten_wert = []

	for i in range(len(spalten)):
		spalten_wert = spalten_wert + [spalten[i].value]

	spalten_wert.remove(name)

	return spalten_wert

#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------

def excel_ausgabe(daten_roh, daten_auswertung_wt_perf, daten_auswertung_wt_perf_jahr, daten_auswertung_wt_span, basiswert):
	
	wb = workbook_erstellen()
	wb = daten_sheet(wb, daten_roh, daten_auswertung_wt_perf, daten_auswertung_wt_perf_jahr, daten_auswertung_wt_span)
	workbook_speichern(wb, basiswert)



def workbook_erstellen():	

	wb = openpyxl.Workbook()
	return wb



def daten_sheet(wb, daten_roh, daten_tage_perf, daten_tage_perf_jahr, daten_tage_span):

	ws = wb.active
	ws.titel = 'Rohdaten'

	da = daten_roh[0]
	op = daten_roh[1]
	high = daten_roh[2]
	low = daten_roh[3]
	clo = daten_roh[4]

	ws.cell(row = 1, column = 1, value = 'Date')
	ws.cell(row = 1, column = 2, value = 'Open')
	ws.cell(row = 1, column = 3, value = 'High')
	ws.cell(row = 1, column = 4, value = 'Low')
	ws.cell(row = 1, column = 5, value = 'Close')

	for i in range(len(daten_roh[0])):
		ws.cell(row = i+3, column = 1, value = da[i])
		ws.cell(row = i+3, column = 2, value = op[i])
		ws.cell(row = i+3, column = 3, value = low[i])
		ws.cell(row = i+3, column = 4, value = high[i])
		ws.cell(row = i+3, column = 5, value = clo[i])

	wb.create_sheet('Wochentage_Performance')
	ws = wb['Wochentage_Performance']

	ws.cell(row = 1 , column = 1, value = 'Wochentage_Performance')
	
	ds = 1 

	ws.cell(row = 2 + ds , column = 1, value = 'Allgemein')
	ws.cell(row = 3 + ds , column = 1, value = 'Anzahl')

	ws.cell(row = 4 + ds, column = 2, value = 'Montag')
	ws.cell(row = 4 + ds, column = 6, value = 'Dienstag')
	ws.cell(row = 4 + ds, column = 10, value = 'Mittwoch')
	ws.cell(row = 4 + ds, column = 14, value = 'Donnerstag')
	ws.cell(row = 4 + ds, column = 18, value = 'Freitag')
	
	for i in range(5):
		ws.cell(row = 5 + ds, column = 2+i*4, value = '9:00 - 17:30')
		ws.cell(row = 5 + ds, column = 3+i*4, value = '17:30 - 9:00')
		ws.cell(row = 5 + ds, column = 4+i*4, value = '9:00 - 9:00')

	ws.cell(row = 6 + ds, column = 1, value = 'Gesamt')
	ws.cell(row = 7 + ds, column = 1, value = 'Positiv')
	ws.cell(row = 8 + ds, column = 1, value = 'Negativ')

	for i in range(5):
		for j in range(3):
			ws.cell(row = 6 + ds, column = 2+i*4+j, value = daten_tage_perf[3][i][j][0])
			ws.cell(row = 7 + ds, column = 2+i*4+j, value = daten_tage_perf[3][i][j][1])
			ws.cell(row = 8 + ds, column = 2+i*4+j, value = daten_tage_perf[3][i][j][2])

	ds += 7

	ws.cell(row = 3 + ds , column = 1, value = 'Durchschnitt (arithmetisch)')

	ws.cell(row = 4 + ds, column = 2, value = 'Montag')
	ws.cell(row = 4 + ds, column = 6, value = 'Dienstag')
	ws.cell(row = 4 + ds, column = 10, value = 'Mittwoch')
	ws.cell(row = 4 + ds, column = 14, value = 'Donnerstag')
	ws.cell(row = 4 + ds, column = 18, value = 'Freitag')

	for i in range(5):
		ws.cell(row = 5 + ds, column = 2+i*4, value = '9:00 - 17:30')
		ws.cell(row = 5 + ds, column = 3+i*4, value = '17:30 - 9:00')
		ws.cell(row = 5 + ds, column = 4+i*4, value = '9:00 - 9:00')

	ws.cell(row = 6 + ds, column = 1, value = 'Absolut')
	ws.cell(row = 7 + ds, column = 1, value = 'Diskrete Rendite')
	ws.cell(row = 8 + ds, column = 1, value = 'Stetige Rendite')

	for i in range(5):
		for j in range(3):
			ws.cell(row = 6 + ds, column = 2+i*4+j, value = daten_tage_perf[1][i][j][0])
			ws.cell(row = 7 + ds, column = 2+i*4+j, value = daten_tage_perf[1][i][j][1])
			ws.cell(row = 8 + ds, column = 2+i*4+j, value = daten_tage_perf[1][i][j][2])

	ds += 7

	ws.cell(row = 3 + ds, column = 1, value = 'Durchschnitt (geometrisch)')

	ws.cell(row = 4 + ds, column = 2, value = 'Montag')
	ws.cell(row = 4 + ds, column = 6, value = 'Dienstag')
	ws.cell(row = 4 + ds, column = 10, value = 'Mittwoch')
	ws.cell(row = 4 + ds, column = 14, value = 'Donnerstag')
	ws.cell(row = 4 + ds, column = 18, value = 'Freitag')

	for i in range(5):
		ws.cell(row = 5 + ds, column = 2+i*4, value = '9:00 - 17:30')
		ws.cell(row = 5 + ds, column = 3+i*4, value = '17:30 - 9:00')
		ws.cell(row = 5 + ds, column = 4+i*4, value = '9:00 - 9:00')

	ws.cell(row = 6 + ds, column = 1, value = 'Diskrete Renditen (1+stetig)')
	
	for i in range(5):
		for j in range(3):
			ws.cell(row = 6 + ds, column = 2+i*4+j, value = daten_tage_perf[2][i][j])

	ds += 9

	ws.cell(row = -1 + ds, column = 1, value = 'Datenreihen')
	ws.cell(row = 1 + ds, column = 1, value = 'Montag')
	ws.cell(row = 1 + ds, column = 18, value = 'Dienstag')
	ws.cell(row = 1 + ds, column = 35, value = 'Mittwoch')
	ws.cell(row = 1 + ds, column = 52, value = 'Donnerstag')
	ws.cell(row = 1 + ds, column = 69, value = 'Freitag')

	for i in range(5):
		ws.cell(row = 2 + ds, column = 1+17*i, value = 'Datum')
		for k in range(len(daten_tage_perf[0][i][0])):
			ws.cell(row = 5 + k + ds, column = 1+17*i, value = daten_tage_perf[0][i][3][k])
		ws.cell(row = 2 + ds, column = 3+17*i, value = '9:00 - 17:30')
		ws.cell(row = 2 + ds, column = 8+17*i, value = '17:30 - 9:00')
		ws.cell(row = 2 + ds, column = 13+17*i, value = '9:00 - 9:00')
		for j in range(3):
			ws.cell(row = 3 + ds, column = 3+5*j+17*i, value = 'Absolut')
			ws.cell(row = 3 + ds, column = 4+5*j+17*i, value = 'Prozent')
			ws.cell(row = 3 + ds, column = 5+5*j+17*i, value = 'Stetig')
			ws.cell(row = 3 + ds, column = 6+5*j+17*i, value = 'Index')
			ws.cell(row = 4 + ds, column = 6+5*j+17*i, value = 100)

			for k in range(len(daten_tage_perf[0][i][0])):
				ws.cell(row = 5 + k + ds, column = 3+5*j+17*i, value = daten_tage_perf[0][i][j][k][0])
				ws.cell(row = 5 + k + ds, column = 4+5*j+17*i, value = daten_tage_perf[0][i][j][k][1])
				ws.cell(row = 5 + k + ds, column = 5+5*j+17*i, value = daten_tage_perf[0][i][j][k][2])
				ws.cell(row = 5 + k + ds, column = 6+5*j+17*i, value = daten_tage_perf[0][i][j][k][3])



	wb.create_sheet('Wochentage_Performance_jährlich')
	ws = wb['Wochentage_Performance_jährlich']

	ws.cell(row = 1 , column = 1, value = 'Wochentage_Performance_jährlich')
	
	ds = 2 
	for h in range(len(daten_tage_perf_jahr[1])):
		ws.cell(row = 1 + ds , column = 1, value = daten_tage_perf_jahr[4]+h)
		ws.cell(row = 2 + ds , column = 1, value = 'Allgemein')
		ws.cell(row = 3 + ds , column = 1, value = 'Anzahl')

		ws.cell(row = 4 + ds, column = 2, value = 'Montag')
		ws.cell(row = 4 + ds, column = 6, value = 'Dienstag')
		ws.cell(row = 4 + ds, column = 10, value = 'Mittwoch')
		ws.cell(row = 4 + ds, column = 14, value = 'Donnerstag')
		ws.cell(row = 4 + ds, column = 18, value = 'Freitag')
		
		for i in range(5):
			ws.cell(row = 5 + ds, column = 2+i*4, value = '9:00 - 17:30')
			ws.cell(row = 5 + ds, column = 3+i*4, value = '17:30 - 9:00')
			ws.cell(row = 5 + ds, column = 4+i*4, value = '9:00 - 9:00')

		ws.cell(row = 6 + ds, column = 1, value = 'Gesamt')
		ws.cell(row = 7 + ds, column = 1, value = 'Positiv')
		ws.cell(row = 8 + ds, column = 1, value = 'Negativ')

		for i in range(5):
			for j in range(3):
				ws.cell(row = 6 + ds, column = 2+i*4+j, value = daten_tage_perf_jahr[3][h][i][j][0])
				ws.cell(row = 7 + ds, column = 2+i*4+j, value = daten_tage_perf_jahr[3][h][i][j][1])
				ws.cell(row = 8 + ds, column = 2+i*4+j, value = daten_tage_perf_jahr[3][h][i][j][2])

		ds += 7

		ws.cell(row = 3 + ds , column = 1, value = 'Durchschnitt (arithmetisch)')

		ws.cell(row = 4 + ds, column = 2, value = 'Montag')
		ws.cell(row = 4 + ds, column = 6, value = 'Dienstag')
		ws.cell(row = 4 + ds, column = 10, value = 'Mittwoch')
		ws.cell(row = 4 + ds, column = 14, value = 'Donnerstag')
		ws.cell(row = 4 + ds, column = 18, value = 'Freitag')

		for i in range(5):
			ws.cell(row = 5 + ds, column = 2+i*4, value = '9:00 - 17:30')
			ws.cell(row = 5 + ds, column = 3+i*4, value = '17:30 - 9:00')
			ws.cell(row = 5 + ds, column = 4+i*4, value = '9:00 - 9:00')

		ws.cell(row = 6 + ds, column = 1, value = 'Absolut')
		ws.cell(row = 7 + ds, column = 1, value = 'Diskrete Rendite')
		ws.cell(row = 8 + ds, column = 1, value = 'Stetige Rendite')

		for i in range(5):
			for j in range(3):
				ws.cell(row = 6 + ds, column = 2+i*4+j, value = daten_tage_perf_jahr[1][h][i][j][0])
				ws.cell(row = 7 + ds, column = 2+i*4+j, value = daten_tage_perf_jahr[1][h][i][j][1])
				ws.cell(row = 8 + ds, column = 2+i*4+j, value = daten_tage_perf_jahr[1][h][i][j][2])

		ds += 7

		ws.cell(row = 3 + ds, column = 1, value = 'Durchschnitt (geometrisch)')

		ws.cell(row = 4 + ds, column = 2, value = 'Montag')
		ws.cell(row = 4 + ds, column = 6, value = 'Dienstag')
		ws.cell(row = 4 + ds, column = 10, value = 'Mittwoch')
		ws.cell(row = 4 + ds, column = 14, value = 'Donnerstag')
		ws.cell(row = 4 + ds, column = 18, value = 'Freitag')

		for i in range(5):
			ws.cell(row = 5 + ds, column = 2+i*4, value = '9:00 - 17:30')
			ws.cell(row = 5 + ds, column = 3+i*4, value = '17:30 - 9:00')
			ws.cell(row = 5 + ds, column = 4+i*4, value = '9:00 - 9:00')

		ws.cell(row = 6 + ds, column = 1, value = 'Diskrete Renditen (1+stetig)')
		
		for i in range(5):
			for j in range(3):
				ws.cell(row = 6 + ds, column = 2+i*4+j, value = daten_tage_perf_jahr[2][h][i][j])

		ds += 9

	for h in range(len(daten_tage_perf_jahr[1])):
		ws.cell(row = -1 + ds , column = 1, value = daten_tage_perf_jahr[4]+h)
		ws.cell(row = 0 + ds, column = 1, value = 'Datenreihen')
		ws.cell(row = 1 + ds, column = 1, value = 'Montag')
		ws.cell(row = 1 + ds, column = 18, value = 'Dienstag')
		ws.cell(row = 1 + ds, column = 35, value = 'Mittwoch')
		ws.cell(row = 1 + ds, column = 52, value = 'Donnerstag')
		ws.cell(row = 1 + ds, column = 69, value = 'Freitag')

		for i in range(5):
			ws.cell(row = 2 + ds, column = 1+17*i, value = 'Datum')
			for k in range(len(daten_tage_perf_jahr[0][h][i][0])):
				ws.cell(row = 5 + k + ds, column = 1+17*i, value = daten_tage_perf_jahr[0][h][i][3][k])
			ws.cell(row = 2 + ds, column = 3+17*i, value = '9:00 - 17:30')
			ws.cell(row = 2 + ds, column = 8+17*i, value = '17:30 - 9:00')
			ws.cell(row = 2 + ds, column = 13+17*i, value = '9:00 - 9:00')
			for j in range(3):
				ws.cell(row = 3 + ds, column = 3+5*j+17*i, value = 'Absolut')
				ws.cell(row = 3 + ds, column = 4+5*j+17*i, value = 'Prozent')
				ws.cell(row = 3 + ds, column = 5+5*j+17*i, value = 'Stetig')
				ws.cell(row = 3 + ds, column = 6+5*j+17*i, value = 'Index')
				ws.cell(row = 4 + ds, column = 6+5*j+17*i, value = 100)

				for k in range(len(daten_tage_perf_jahr[0][h][i][0])):
					ws.cell(row = 5 + k + ds, column = 3+5*j+17*i, value = daten_tage_perf_jahr[0][h][i][j][k][0])
					ws.cell(row = 5 + k + ds, column = 4+5*j+17*i, value = daten_tage_perf_jahr[0][h][i][j][k][1])
					ws.cell(row = 5 + k + ds, column = 5+5*j+17*i, value = daten_tage_perf_jahr[0][h][i][j][k][2])
					ws.cell(row = 5 + k + ds, column = 6+5*j+17*i, value = daten_tage_perf_jahr[0][h][i][j][k][3])

		ds += 12
		ds += len(daten_tage_perf_jahr[0][h][i][0])

	wb.create_sheet('Wochentage_Spanne')
	ws = wb['Wochentage_Spanne']

	ws.cell(row = 1 , column = 1, value = 'Wochentage_Spanne')
	ws.cell(row = 3 , column = 1, value = 'Durchschnitt (arithmetisch)')

	ws.cell(row = 4, column = 2, value = 'Montag')
	ws.cell(row = 4, column = 3, value = 'Dienstag')
	ws.cell(row = 4, column = 4, value = 'Mittwoch')
	ws.cell(row = 4, column = 5, value = 'Donnerstag')
	ws.cell(row = 4, column = 6, value = 'Freitag')

	ws.cell(row = 5 , column = 1, value = 'Absolut')
	ws.cell(row = 6 , column = 1, value = 'Prozent')

	for i in range(5):
		ws.cell(row = 5 , column = 2+i, value = daten_tage_span[1][i][0])
		ws.cell(row = 6 , column = 2+i, value = daten_tage_span[1][i][1])		

	ds = 9

	ws.cell(row = 0 + ds, column = 1, value = 'Datenreihen')
	ws.cell(row = 1 + ds, column = 1, value = 'Montag')
	ws.cell(row = 1 + ds, column = 6, value = 'Dienstag')
	ws.cell(row = 1 + ds, column = 11, value = 'Mittwoch')
	ws.cell(row = 1 + ds, column = 16, value = 'Donnerstag')
	ws.cell(row = 1 + ds, column = 21, value = 'Freitag')

	for i in range(5):
		ws.cell(row = 2 + ds, column = 1+5*i, value = 'Datum')
		for k in range(len(daten_tage_span[0][i][2])):
			ws.cell(row = 3 + k + ds, column = 1+5*i, value = daten_tage_span[0][i][2][k])
		ws.cell(row = 2 + ds, column = 3+5*i, value = 'Absolut')
		ws.cell(row = 2 + ds, column = 4+5*i, value = 'Prozent')
		for k in range(len(daten_tage_span[0][i][2])):
			ws.cell(row = 3 + k + ds, column = 3+5*i, value = daten_tage_span[0][i][0][k])
			ws.cell(row = 3 + k + ds, column = 4+5*i, value = daten_tage_span[0][i][1][k])

	return wb

def workbook_speichern(w, basiswert):
	wb = w
	pfad1 = 'Output'
	pfad2 = basiswert + '_Wochentage_Monate'
	verzeichnis = os.path.join(pfad1,pfad2)
	
	if not os.path.exists(verzeichnis):
		os.makedirs(verzeichnis)

	wb.save(verzeichnis +'\ ' + basiswert + '_Wochentage_Monate.xlsx')

#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------------------------------------

def Wochentage_Performance(datum, ope, close):

	year = []
	
	for i in range(len(datum)):
		dat = datum[i]
		year += [dat.year]

	min_year = min(year)
	max_year = max(year)

	daten_aufbe = []

	for i in range(max_year-min_year+1):
		daten_aufbe += [[min_year+i,[]]]

	for i in range(len(datum)):
		dat = datum[i]
		kerzen = [ope[i],close[i]] 
		for j in range(max_year-min_year+1):
			if dat.year == min_year + j:
				daten_aufbe[j][1] += [[dat.weekday(), dat.isocalendar()[1], dat.month, kerzen]]

	wochentage = []
	wochentage_anzahl = []

	for i in range(5):
		wochentage += [[[],[],[],[]]]
		wochentage_anzahl += [[[0,0,0],[0,0,0],[0,0,0]]]

	m = 0
	index = 100

	for i in range(len(daten_aufbe)):
		for j in range(len(daten_aufbe[i][1])):

			tag_9_17_abs = (daten_aufbe[i][1][j][3][1] - daten_aufbe[i][1][j][3][0])  
			tag_9_17_pro = tag_9_17_abs / daten_aufbe[i][1][j][3][0]
			tag_9_17_stet = np.log(daten_aufbe[i][1][j][3][1]) - np.log(daten_aufbe[i][1][j][3][0])
			
			if len(wochentage[daten_aufbe[i][1][j][0]][0]) == 0:
				tag_9_17_index = index*np.exp(tag_9_17_stet)
			else:
				index_akt = wochentage[daten_aufbe[i][1][j][0]][0][len(wochentage[daten_aufbe[i][1][j][0]][0])-1][3]
				tag_9_17_index = index_akt*np.exp(tag_9_17_stet)
			
			tag_9_17 = [tag_9_17_abs, tag_9_17_pro, tag_9_17_stet, tag_9_17_index]



			if i < len(daten_aufbe)-1:
				if j < len(daten_aufbe[i][1])-1:
					tag_17_9_abs = (daten_aufbe[i][1][j+1][3][0] - daten_aufbe[i][1][j][3][1])  
					tag_17_9_pro = tag_17_9_abs / daten_aufbe[i][1][j][3][1]
					tag_17_9_stet = np.log(daten_aufbe[i][1][j+1][3][0]) - np.log(daten_aufbe[i][1][j][3][1])
								
					if len(wochentage[daten_aufbe[i][1][j][0]][1]) == 0:
						tag_17_9_index = index*np.exp(tag_17_9_stet)
					else:
						index_akt = wochentage[daten_aufbe[i][1][j][0]][1][len(wochentage[daten_aufbe[i][1][j][0]][1])-1][3]
						tag_17_9_index = index_akt*np.exp(tag_17_9_stet)
	
					tag_17_9 = [tag_17_9_abs, tag_17_9_pro, tag_17_9_stet, tag_17_9_index]
				elif j == len(daten_aufbe[i][1])-1:
					tag_17_9_abs = (daten_aufbe[i+1][1][0][3][0] - daten_aufbe[i][1][j][3][1])  
					tag_17_9_pro = tag_17_9_abs / daten_aufbe[i][1][j][3][1]
					tag_17_9_stet = np.log(daten_aufbe[i+1][1][0][3][0]) - np.log(daten_aufbe[i][1][j][3][1])					
			
					if len(wochentage[daten_aufbe[i][1][j][0]][1]) == 0:
						tag_17_9_index = index*np.exp(tag_17_9_stet)
					else:
						index_akt = wochentage[daten_aufbe[i][1][j][0]][1][len(wochentage[daten_aufbe[i][1][j][0]][1])-1][3]
						tag_17_9_index = index_akt*np.exp(tag_17_9_stet)
	
					tag_17_9 = [tag_17_9_abs, tag_17_9_pro, tag_17_9_stet, tag_17_9_index]
			elif i == len(daten_aufbe)-1:
				if j < len(daten_aufbe[i][1])-1:
					tag_17_9_abs = (daten_aufbe[i][1][j+1][3][0] - daten_aufbe[i][1][j][3][1])  
					tag_17_9_pro = tag_17_9_abs / daten_aufbe[i][1][j][3][1]
					tag_17_9_stet = np.log(daten_aufbe[i][1][j+1][3][0]) - np.log(daten_aufbe[i][1][j][3][1])

					if len(wochentage[daten_aufbe[i][1][j][0]][1]) == 0:
						tag_17_9_index = index*np.exp(tag_17_9_stet)
					else:
						index_akt = wochentage[daten_aufbe[i][1][j][0]][1][len(wochentage[daten_aufbe[i][1][j][0]][1])-1][3]
						tag_17_9_index = index_akt*np.exp(tag_17_9_stet)
	
					tag_17_9 = [tag_17_9_abs, tag_17_9_pro, tag_17_9_stet, tag_17_9_index]
				if j == len(daten_aufbe[i][1])-1:
					tag_17_9 = [0, 0, 0, 0]

			if i < len(daten_aufbe)-1:
				if j < len(daten_aufbe[i][1])-1:
					tag_9_9_abs = (daten_aufbe[i][1][j+1][3][0] - daten_aufbe[i][1][j][3][0])  
					tag_9_9_pro = tag_9_9_abs / daten_aufbe[i][1][j][3][0]
					tag_9_9_stet = np.log(daten_aufbe[i][1][j+1][3][0]) - np.log(daten_aufbe[i][1][j][3][0])

					if len(wochentage[daten_aufbe[i][1][j][0]][2]) == 0:
						tag_9_9_index = index*np.exp(tag_9_9_stet)
					else:
						index_akt = wochentage[daten_aufbe[i][1][j][0]][2][len(wochentage[daten_aufbe[i][1][j][0]][2])-1][3]
						tag_9_9_index = index_akt*np.exp(tag_9_9_stet)

					tag_9_9 = [tag_9_9_abs, tag_9_9_pro, tag_9_9_stet, tag_9_9_index]

				elif j == len(daten_aufbe[i][1])-1:
					tag_9_9_abs = (daten_aufbe[i+1][1][0][3][0] - daten_aufbe[i][1][j][3][0])  
					tag_9_9_pro = tag_9_9_abs / daten_aufbe[i][1][j][3][0]
					tag_9_9_stet = np.log(daten_aufbe[i+1][1][0][3][0]) - np.log(daten_aufbe[i][1][j][3][0])

					if len(wochentage[daten_aufbe[i][1][j][0]][2]) == 0:
						tag_9_9_index = index*np.exp(tag_9_9_stet)
					else:
						index_akt = wochentage[daten_aufbe[i][1][j][0]][2][len(wochentage[daten_aufbe[i][1][j][0]][2])-1][3]
						tag_9_9_index = index_akt*np.exp(tag_9_9_stet)

					tag_9_9 = [tag_9_9_abs, tag_9_9_pro, tag_9_9_stet, tag_9_9_index]

			elif i == len(daten_aufbe)-1:
				if j < len(daten_aufbe[i][1])-1:
					tag_9_9_abs = (daten_aufbe[i][1][j+1][3][0] - daten_aufbe[i][1][j][3][0])  
					tag_9_9_pro = tag_9_9_abs / daten_aufbe[i][1][j][3][0]
					tag_9_9_stet = np.log(daten_aufbe[i][1][j+1][3][0]) - np.log(daten_aufbe[i][1][j][3][0])

					if len(wochentage[daten_aufbe[i][1][j][0]][2]) == 0:
						tag_9_9_index = index*np.exp(tag_9_9_stet)
					else:
						index_akt = wochentage[daten_aufbe[i][1][j][0]][2][len(wochentage[daten_aufbe[i][1][j][0]][2])-1][3]
						tag_9_9_index = index_akt*np.exp(tag_9_9_stet)

					tag_9_9 = [tag_9_9_abs, tag_9_9_pro, tag_9_9_stet, tag_9_9_index]
				if j == len(daten_aufbe[i][1])-1:
					tag_9_9 = [0, 0, 0, 0]



			if daten_aufbe[i][1][j][0] == 0:
				wochentage[0][0] += [tag_9_17]
				wochentage[0][1] += [tag_17_9]
				wochentage[0][2] += [tag_9_9]
				wochentage[0][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[0][0][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[0][0][2] += 1
					else:
						wochentage_anzahl[0][0][1] += 1
				else:
					wochentage_anzahl[0][0][0] += 1
					wochentage_anzahl[0][1][0] += 1
					wochentage_anzahl[0][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[0][0][2] += 1
					else:
						wochentage_anzahl[0][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[0][1][2] += 1
					else:
						wochentage_anzahl[0][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[0][2][2] += 1
					else:
						wochentage_anzahl[0][2][1] += 1
			elif daten_aufbe[i][1][j][0] == 1:
				wochentage[1][0] += [tag_9_17]
				wochentage[1][1] += [tag_17_9]
				wochentage[1][2] += [tag_9_9]
				wochentage[1][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[1][0][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[1][0][2] += 1
					else:
						wochentage_anzahl[1][0][1] += 1
				else:
					wochentage_anzahl[1][0][0] += 1
					wochentage_anzahl[1][1][0] += 1
					wochentage_anzahl[1][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[1][0][2] += 1
					else:
						wochentage_anzahl[1][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[1][1][2] += 1
					else:
						wochentage_anzahl[1][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[1][2][2] += 1
					else:
						wochentage_anzahl[1][2][1] += 1
			elif daten_aufbe[i][1][j][0] == 2:
				wochentage[2][0] += [tag_9_17]
				wochentage[2][1] += [tag_17_9]
				wochentage[2][2] += [tag_9_9]
				wochentage[2][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[2][0][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[2][0][2] += 1
					else:
						wochentage_anzahl[2][0][1] += 1
				else:
					wochentage_anzahl[2][0][0] += 1
					wochentage_anzahl[2][1][0] += 1
					wochentage_anzahl[2][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[2][0][2] += 1
					else:
						wochentage_anzahl[2][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[2][1][2] += 1
					else:
						wochentage_anzahl[2][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[2][2][2] += 1
					else:
						wochentage_anzahl[2][2][1] += 1
			elif daten_aufbe[i][1][j][0] == 3:
				wochentage[3][0] += [tag_9_17]
				wochentage[3][1] += [tag_17_9]
				wochentage[3][2] += [tag_9_9]
				wochentage[3][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[0][3][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[3][0][2] += 1
					else:
						wochentage_anzahl[3][0][1] += 1
				else:
					wochentage_anzahl[3][0][0] += 1
					wochentage_anzahl[3][1][0] += 1
					wochentage_anzahl[3][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[3][0][2] += 1
					else:
						wochentage_anzahl[3][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[3][1][2] += 1
					else:
						wochentage_anzahl[3][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[3][2][2] += 1
					else:
						wochentage_anzahl[3][2][1] += 1
			elif daten_aufbe[i][1][j][0] == 4:
				wochentage[4][0] += [tag_9_17]
				wochentage[4][1] += [tag_17_9]
				wochentage[4][2] += [tag_9_9]
				wochentage[4][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[4][0][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[4][0][2] += 1
					else:
						wochentage_anzahl[4][0][1] += 1
				else:
					wochentage_anzahl[4][0][0] += 1
					wochentage_anzahl[4][1][0] += 1
					wochentage_anzahl[4][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[4][0][2] += 1
					else:
						wochentage_anzahl[4][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[4][1][2] += 1
					else:
						wochentage_anzahl[4][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[4][2][2] += 1
					else:
						wochentage_anzahl[4][2][1] += 1

			m += 1

	durchschnitt_wochentage =[]

	for i in range(5):
		durchschnitt_wochentage += [[[],[],[]]]

	for i in range(5):
		for j in range(3):

			gesamt_punkte = 0
			gesamt_pro = 0
			gesamt_stet = 0

			for k in range(len(wochentage[i][j])):
				gesamt_punkte += wochentage[i][j][k][0]
				gesamt_pro += wochentage[i][j][k][1]
				gesamt_stet += wochentage[i][j][k][2]

			durchschnitt_punkte = gesamt_punkte / len(wochentage[i][j])
			durchschnitt_pro = gesamt_pro / len(wochentage[i][j])
			durchschnitt_stet = gesamt_stet / len(wochentage[i][j])

			durchschnitt_wochentage[i][j] = [durchschnitt_punkte, durchschnitt_pro, durchschnitt_stet, len(wochentage[i][j])]

	geo_durchschnitt_wt = []

	for i in range(5):
		geo_durchschnitt_wt += [[0,0,0]]

	for i in range(5):
		for j in range(3):

			gesamt_stet = []
			for k in range(len(wochentage[i][j])):
				gesamt_stet += [wochentage[i][j][k][1]+1]
			geo_durchschnitt_wt[i][j] = scistats.gmean(gesamt_stet)	

	return [wochentage, durchschnitt_wochentage, geo_durchschnitt_wt, wochentage_anzahl]

def Wochentage_Performance_Year(datum, ope, close):

	year = []
		
	for i in range(len(datum)):
		dat = datum[i]
		year += [dat.year]

	min_year = min(year)
	max_year = max(year)

	daten_aufbe = []

	for i in range(max_year-min_year+1):
		daten_aufbe += [[min_year+i,[]]]

	for i in range(len(datum)):
		dat = datum[i]
		kerzen = [ope[i],close[i]] 
		for j in range(max_year-min_year+1):
			if dat.year == min_year + j:
				daten_aufbe[j][1] += [[dat.weekday(), dat.isocalendar()[1], dat.month, kerzen]]

	wochentage = []
	wochentage_anzahl = []
	for i in range(max_year-min_year+1):
		wochentage += [[]]
		wochentage_anzahl += [[]]
		for j in range(5):
			wochentage[i] += [[[],[],[],[]]]
			wochentage_anzahl[i] += [[[0,0,0],[0,0,0],[0,0,0]]]

	m = 0
	index = 100

	for i in range(len(daten_aufbe)):
		for j in range(len(daten_aufbe[i][1])):

			tag_9_17_abs = (daten_aufbe[i][1][j][3][1] - daten_aufbe[i][1][j][3][0])  
			tag_9_17_pro = tag_9_17_abs / daten_aufbe[i][1][j][3][0]
			tag_9_17_stet = np.log(daten_aufbe[i][1][j][3][1]) - np.log(daten_aufbe[i][1][j][3][0])
			
			if len(wochentage[i][daten_aufbe[i][1][j][0]][0]) == 0:
				tag_9_17_index = index*np.exp(tag_9_17_stet)
			else:
				index_akt = wochentage[i][daten_aufbe[i][1][j][0]][0][len(wochentage[i][daten_aufbe[i][1][j][0]][0])-1][3]
				tag_9_17_index = index_akt*np.exp(tag_9_17_stet)
			
			tag_9_17 = [tag_9_17_abs, tag_9_17_pro, tag_9_17_stet, tag_9_17_index]



			if i < len(daten_aufbe)-1:
				if j < len(daten_aufbe[i][1])-1:
					tag_17_9_abs = (daten_aufbe[i][1][j+1][3][0] - daten_aufbe[i][1][j][3][1])  
					tag_17_9_pro = tag_17_9_abs / daten_aufbe[i][1][j][3][1]
					tag_17_9_stet = np.log(daten_aufbe[i][1][j+1][3][0]) - np.log(daten_aufbe[i][1][j][3][1])
								
					if len(wochentage[i][daten_aufbe[i][1][j][0]][1]) == 0:
						tag_17_9_index = index*np.exp(tag_17_9_stet)
					else:
						index_akt = wochentage[i][daten_aufbe[i][1][j][0]][1][len(wochentage[i][daten_aufbe[i][1][j][0]][1])-1][3]
						tag_17_9_index = index_akt*np.exp(tag_17_9_stet)
	
					tag_17_9 = [tag_17_9_abs, tag_17_9_pro, tag_17_9_stet, tag_17_9_index]
				elif j == len(daten_aufbe[i][1])-1:
					tag_17_9_abs = (daten_aufbe[i+1][1][0][3][0] - daten_aufbe[i][1][j][3][1])  
					tag_17_9_pro = tag_17_9_abs / daten_aufbe[i][1][j][3][1]
					tag_17_9_stet = np.log(daten_aufbe[i+1][1][0][3][0]) - np.log(daten_aufbe[i][1][j][3][1])					
			
					if len(wochentage[i][daten_aufbe[i][1][j][0]][1]) == 0:
						tag_17_9_index = index*np.exp(tag_17_9_stet)
					else:
						index_akt = wochentage[i][daten_aufbe[i][1][j][0]][1][len(wochentage[i][daten_aufbe[i][1][j][0]][1])-1][3]
						tag_17_9_index = index_akt*np.exp(tag_17_9_stet)
	
					tag_17_9 = [tag_17_9_abs, tag_17_9_pro, tag_17_9_stet, tag_17_9_index]
			elif i == len(daten_aufbe)-1:
				if j < len(daten_aufbe[i][1])-1:
					tag_17_9_abs = (daten_aufbe[i][1][j+1][3][0] - daten_aufbe[i][1][j][3][1])  
					tag_17_9_pro = tag_17_9_abs / daten_aufbe[i][1][j][3][1]
					tag_17_9_stet = np.log(daten_aufbe[i][1][j+1][3][0]) - np.log(daten_aufbe[i][1][j][3][1])

					if len(wochentage[i][daten_aufbe[i][1][j][0]][1]) == 0:
						tag_17_9_index = index*np.exp(tag_17_9_stet)
					else:
						index_akt = wochentage[i][daten_aufbe[i][1][j][0]][1][len(wochentage[i][daten_aufbe[i][1][j][0]][1])-1][3]
						tag_17_9_index = index_akt*np.exp(tag_17_9_stet)
	
					tag_17_9 = [tag_17_9_abs, tag_17_9_pro, tag_17_9_stet, tag_17_9_index]
				if j == len(daten_aufbe[i][1])-1:
					tag_17_9 = [0, 0, 0, 0]

			if i < len(daten_aufbe)-1:
				if j < len(daten_aufbe[i][1])-1:
					tag_9_9_abs = (daten_aufbe[i][1][j+1][3][0] - daten_aufbe[i][1][j][3][0])  
					tag_9_9_pro = tag_9_9_abs / daten_aufbe[i][1][j][3][0]
					tag_9_9_stet = np.log(daten_aufbe[i][1][j+1][3][0]) - np.log(daten_aufbe[i][1][j][3][0])

					if len(wochentage[i][daten_aufbe[i][1][j][0]][2]) == 0:
						tag_9_9_index = index*np.exp(tag_9_9_stet)
					else:
						index_akt = wochentage[i][daten_aufbe[i][1][j][0]][2][len(wochentage[i][daten_aufbe[i][1][j][0]][2])-1][3]
						tag_9_9_index = index_akt*np.exp(tag_9_9_stet)

					tag_9_9 = [tag_9_9_abs, tag_9_9_pro, tag_9_9_stet, tag_9_9_index]

				elif j == len(daten_aufbe[i][1])-1:
					tag_9_9_abs = (daten_aufbe[i+1][1][0][3][0] - daten_aufbe[i][1][j][3][0])  
					tag_9_9_pro = tag_9_9_abs / daten_aufbe[i][1][j][3][0]
					tag_9_9_stet = np.log(daten_aufbe[i+1][1][0][3][0]) - np.log(daten_aufbe[i][1][j][3][0])

					if len(wochentage[i][daten_aufbe[i][1][j][0]][2]) == 0:
						tag_9_9_index = index*np.exp(tag_9_9_stet)
					else:
						index_akt = wochentage[i][daten_aufbe[i][1][j][0]][2][len(wochentage[i][daten_aufbe[i][1][j][0]][2])-1][3]
						tag_9_9_index = index_akt*np.exp(tag_9_9_stet)

					tag_9_9 = [tag_9_9_abs, tag_9_9_pro, tag_9_9_stet, tag_9_9_index]

			elif i == len(daten_aufbe)-1:
				if j < len(daten_aufbe[i][1])-1:
					tag_9_9_abs = (daten_aufbe[i][1][j+1][3][0] - daten_aufbe[i][1][j][3][0])  
					tag_9_9_pro = tag_9_9_abs / daten_aufbe[i][1][j][3][0]
					tag_9_9_stet = np.log(daten_aufbe[i][1][j+1][3][0]) - np.log(daten_aufbe[i][1][j][3][0])

					if len(wochentage[i][daten_aufbe[i][1][j][0]][2]) == 0:
						tag_9_9_index = index*np.exp(tag_9_9_stet)
					else:
						index_akt = wochentage[i][daten_aufbe[i][1][j][0]][2][len(wochentage[i][daten_aufbe[i][1][j][0]][2])-1][3]
						tag_9_9_index = index_akt*np.exp(tag_9_9_stet)

					tag_9_9 = [tag_9_9_abs, tag_9_9_pro, tag_9_9_stet, tag_9_9_index]
				if j == len(daten_aufbe[i][1])-1:
					tag_9_9 = [0, 0, 0, 0]
	

			if daten_aufbe[i][1][j][0] == 0:
				wochentage[i][0][0] += [tag_9_17]
				wochentage[i][0][1] += [tag_17_9]
				wochentage[i][0][2] += [tag_9_9]
				wochentage[i][0][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[i][0][0][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][0][0][2] += 1
					else:
						wochentage_anzahl[i][0][0][1] += 1
				else:
					wochentage_anzahl[i][0][0][0] += 1
					wochentage_anzahl[i][0][1][0] += 1
					wochentage_anzahl[i][0][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][0][0][2] += 1
					else:
						wochentage_anzahl[i][0][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[i][0][1][2] += 1
					else:
						wochentage_anzahl[i][0][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[i][0][2][2] += 1
					else:
						wochentage_anzahl[i][0][2][1] += 1
			elif daten_aufbe[i][1][j][0] == 1:
				wochentage[i][1][0] += [tag_9_17]
				wochentage[i][1][1] += [tag_17_9]
				wochentage[i][1][2] += [tag_9_9]
				wochentage[i][1][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[i][1][0][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][1][0][2] += 1
					else:
						wochentage_anzahl[i][1][0][1] += 1
				else:
					wochentage_anzahl[i][1][0][0] += 1
					wochentage_anzahl[i][1][1][0] += 1
					wochentage_anzahl[i][1][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][1][0][2] += 1
					else:
						wochentage_anzahl[i][1][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[i][1][1][2] += 1
					else:
						wochentage_anzahl[i][1][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[i][1][2][2] += 1
					else:
						wochentage_anzahl[i][1][2][1] += 1
			elif daten_aufbe[i][1][j][0] == 2:
				wochentage[i][2][0] += [tag_9_17]
				wochentage[i][2][1] += [tag_17_9]
				wochentage[i][2][2] += [tag_9_9]
				wochentage[i][2][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[i][2][0][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][2][0][2] += 1
					else:
						wochentage_anzahl[i][2][0][1] += 1
				else:
					wochentage_anzahl[i][2][0][0] += 1
					wochentage_anzahl[i][2][1][0] += 1
					wochentage_anzahl[i][2][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][2][0][2] += 1
					else:
						wochentage_anzahl[i][2][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[i][2][1][2] += 1
					else:
						wochentage_anzahl[i][2][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[i][2][2][2] += 1
					else:
						wochentage_anzahl[i][2][2][1] += 1
			elif daten_aufbe[i][1][j][0] == 3:
				wochentage[i][3][0] += [tag_9_17]
				wochentage[i][3][1] += [tag_17_9]
				wochentage[i][3][2] += [tag_9_9]
				wochentage[i][3][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[i][0][3][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][3][0][2] += 1
					else:
						wochentage_anzahl[i][3][0][1] += 1
				else:
					wochentage_anzahl[i][3][0][0] += 1
					wochentage_anzahl[i][3][1][0] += 1
					wochentage_anzahl[i][3][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][3][0][2] += 1
					else:
						wochentage_anzahl[i][3][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[i][3][1][2] += 1
					else:
						wochentage_anzahl[i][3][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[i][3][2][2] += 1
					else:
						wochentage_anzahl[i][3][2][1] += 1
			elif daten_aufbe[i][1][j][0] == 4:
				wochentage[i][4][0] += [tag_9_17]
				wochentage[i][4][1] += [tag_17_9]
				wochentage[i][4][2] += [tag_9_9]
				wochentage[i][4][3] += [datum[m]]
				if i == len(daten_aufbe)-1 and j == len(daten_aufbe[i][1])-1:
					wochentage_anzahl[i][4][0][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][4][0][2] += 1
					else:
						wochentage_anzahl[i][4][0][1] += 1
				else:
					wochentage_anzahl[i][4][0][0] += 1
					wochentage_anzahl[i][4][1][0] += 1
					wochentage_anzahl[i][4][2][0] += 1
					if tag_9_17[0] < 0:
						wochentage_anzahl[i][4][0][2] += 1
					else:
						wochentage_anzahl[i][4][0][1] += 1
					if tag_17_9[0] < 0:
						wochentage_anzahl[i][4][1][2] += 1
					else:
						wochentage_anzahl[i][4][1][1] += 1
					if tag_9_9[0] < 0:
						wochentage_anzahl[i][4][2][2] += 1
					else:
						wochentage_anzahl[i][4][2][1] += 1

			m += 1

	durchschnitt_wochentage =[]
	for i in range(max_year-min_year+1):
		durchschnitt_wochentage += [[]]
		for j in range(5):
			durchschnitt_wochentage[i] += [[[],[],[]]]
	for h in range(max_year-min_year+1):
		for i in range(5):
			for j in range(3):

				gesamt_punkte = 0
				gesamt_pro = 0
				gesamt_stet = 0

				for k in range(len(wochentage[h][i][j])):
					gesamt_punkte += wochentage[h][i][j][k][0]
					gesamt_pro += wochentage[h][i][j][k][1]
					gesamt_stet += wochentage[h][i][j][k][2]

				durchschnitt_punkte = gesamt_punkte / len(wochentage[h][i][j])
				durchschnitt_pro = gesamt_pro / len(wochentage[h][i][j])
				durchschnitt_stet = gesamt_stet / len(wochentage[h][i][j])

				durchschnitt_wochentage[h][i][j] = [durchschnitt_punkte, durchschnitt_pro, durchschnitt_stet, len(wochentage[i][j])]

	geo_durchschnitt_wt = []
	for i in range(max_year-min_year+1):
		geo_durchschnitt_wt += [[]]
		for j in range(5):
			geo_durchschnitt_wt[i] += [[0,0,0]]
	for h in range(max_year-min_year+1):
		for i in range(5):
			for j in range(3):

				gesamt_stet = []
				for k in range(len(wochentage[h][i][j])):
					gesamt_stet += [wochentage[h][i][j][k][1]+1]
				geo_durchschnitt_wt[h][i][j] = scistats.gmean(gesamt_stet)	

	return [wochentage, durchschnitt_wochentage, geo_durchschnitt_wt, wochentage_anzahl, min_year]



def Wochentage_Spanne(datum, ope, high, low):

	spanne_wochentage = [[[],[],[]],[[],[],[]],[[],[],[]],[[],[],[]],[[],[],[]]]

	for i in range(len(datum)):

		if datum[i].weekday() == 0:
			spanne_wochentage[0][0] += [high[i]-low[i]]
			spanne_wochentage[0][1] += [(high[i]-low[i])/ope[i]]
			spanne_wochentage[0][2] += [datum[i]]
		elif datum[i].weekday() == 1:
			spanne_wochentage[1][0] += [high[i]-low[i]]
			spanne_wochentage[1][1] += [(high[i]-low[i])/ope[i]]
			spanne_wochentage[1][2] += [datum[i]]
		elif datum[i].weekday() == 2:
			spanne_wochentage[2][0] += [high[i]-low[i]]
			spanne_wochentage[2][1] += [(high[i]-low[i])/ope[i]]
			spanne_wochentage[2][2] += [datum[i]]
		elif datum[i].weekday() == 3:
			spanne_wochentage[3][0] += [high[i]-low[i]]
			spanne_wochentage[3][1] += [(high[i]-low[i])/ope[i]]
			spanne_wochentage[3][2] += [datum[i]]
		elif datum[i].weekday() == 4:
			spanne_wochentage[4][0] += [high[i]-low[i]]
			spanne_wochentage[4][1] += [(high[i]-low[i])/ope[i]]
			spanne_wochentage[4][2] += [datum[i]]

	durchschnitt_spanne_wochentage = []

	for i in range(5):

		durchschnitt_spanne_wochentage += [[np.mean(spanne_wochentage[i][0]), np.mean(spanne_wochentage[i][1])]]

	return[spanne_wochentage, durchschnitt_spanne_wochentage]



def index_wt_aufbereitung(daten):

	index_list = []
	for i in range(5):
		index_list += [[[100],[100],[100]]]

	for i in range(5):
		for j in range(3):
			for k in range(len(daten[0][i][0])):
				if daten[0][i][j][k][3] != 0:
					index_list[i][j] += [daten[0][i][j][k][3]]

	return index_list



def disk_wt_aufbereitung(daten):

	disk_list = []
	for i in range(5):
		disk_list += [[[100],[100],[100]]]

	for i in range(5):
		for j in range(3):
			for k in range(len(daten[0][i][0])):
				if daten[0][i][j][k][1] != 0:
					disk_list[i][j] += [daten[0][i][j][k][1]]

	return disk_list



def stet_wt_aufbereitung(daten):

	stet_list = []
	for i in range(5):
		stet_list += [[[100],[100],[100]]]

	for i in range(5):
		for j in range(3):
			for k in range(len(daten[0][i][0])):
				if daten[0][i][j][k][2] != 0:
					stet_list[i][j] += [daten[0][i][j][k][2]]

	return stet_list

def plot_wt_tage(index,basiswert):

	font = {'fontname':'Calibri'}

	tage_list = ['Montag','Dientag','Mittwoch','Donnerstag','Freitag']
	pfad1 = 'Output'
	pfad2 = basiswert + '_Wochentage_Monate'
	pfad3 = 'Bilder'
	verzeichnis = os.path.join(pfad1,pfad2,pfad3)
	
	if not os.path.exists(verzeichnis):
		os.makedirs(verzeichnis)

	for i in range(5):
		fig, ax = plt.subplots(facecolor='#222222')
		plt.title(tage_list[i], fontweight = 'bold',fontsize = 18, **font, color = 'white')
		ax.plot(range(len(index[i][0])),index[i][0], '#004E8A', linewidth=2.0, label = 'Intraday')
		ax.plot(range(len(index[i][1])),index[i][1], '#B90F22', linewidth=2.0, label = 'Overnight')
		ax.plot(range(len(index[i][2])),index[i][2], '#7FAB16', linewidth=2.0, label = 'Gesamter Tag')
		ax.set(xlabel = 'Anzahl Tage', ylabel = 'Indexstand')
		ax.legend()
		ax.grid()
		ax.set_facecolor('#222222')
		ax.spines['bottom'].set_color('white')
		ax.spines['top'].set_color('white') 
		ax.spines['right'].set_color('white')
		ax.spines['left'].set_color('white')
		ax.tick_params(axis='x', colors='white')
		ax.tick_params(axis='y', colors='white')
		plt.ylabel('Index', **font, color = 'white')
		plt.xlabel('Anzahl', **font, color = 'white')

		plt.savefig(verzeichnis +'\ ' + tage_list[i] + '_gesamt.png', dpi = 900, facecolor=fig.get_facecolor(), edgecolor='none')

	tage_list = ['9_17','17_9','9_9']
	for i in range(3):
		fig, ax = plt.subplots(facecolor='#222222')
		plt.title(tage_list[i], fontweight = 'bold',fontsize = 18, **font, color = 'white')
		ax.plot(range(len(index[0][i])),index[0][i], '#004E8A', linewidth=2.0, label = 'Montag')
		ax.plot(range(len(index[1][i])),index[1][i], '#B90F22', linewidth=2.0, label = 'Dienstag')
		ax.plot(range(len(index[2][i])),index[2][i], '#7FAB16', linewidth=2.0, label = 'Mittwoch')
		ax.plot(range(len(index[3][i])),index[3][i], '#D7AC00', linewidth=2.0, label = 'Donnerstag')
		ax.plot(range(len(index[4][i])),index[4][i], '#611C73', linewidth=2.0, label = 'Freitag')
		ax.legend()
		ax.grid()
		ax.set_facecolor('#222222')
		ax.spines['bottom'].set_color('white')
		ax.spines['top'].set_color('white') 
		ax.spines['right'].set_color('white')
		ax.spines['left'].set_color('white')
		ax.tick_params(axis='x', colors='white')
		ax.tick_params(axis='y', colors='white')
		plt.ylabel('Index', **font, color = 'white')
		plt.xlabel('Anzahl', **font, color = 'white')

		plt.savefig(verzeichnis +'\ ' + tage_list[i] + '_gesamt.png', dpi = 300, facecolor=fig.get_facecolor(), edgecolor='none')

def bar_Wt_tage(durchschnitt,basiswert):

	font = {'fontname':'Calibri'}
	bars_list = ['9_17','17_9','9_9']
	tage_list = ['Montag','Dientag','Mittwoch','Donnerstag','Freitag']
	pfad1 = 'Output'
	pfad2 = basiswert + '_Wochentage_Monate'
	pfad3 = 'Bilder'
	verzeichnis = os.path.join(pfad1,pfad2,pfad3)
	
	if not os.path.exists(verzeichnis):
		os.makedirs(verzeichnis)

	for i in range(3):
		ind = np.arange(5)
		avg_bar = (durchschnitt[1][0][i][2]*100, durchschnitt[1][1][i][2]*100, durchschnitt[1][2][i][2]*100, durchschnitt[1][3][i][2]*100, durchschnitt[1][4][i][2]*100)

		fig, ax = plt.subplots(facecolor='#222222')

		rects = ax.bar(ind, avg_bar, 0.5, color = '#004E8A')

		ax.set_facecolor('#222222')
		ax.spines['bottom'].set_color('white')
		ax.spines['top'].set_color('white') 
		ax.spines['right'].set_color('white')
		ax.spines['left'].set_color('white')
		ax.tick_params(axis='x', colors='white')
		ax.tick_params(axis='y', colors='white')

		plt.title('Durchschnitt stetiger Renditen', fontweight = 'bold',fontsize = 18, **font, color = 'white')
		plt.ylabel('Prozent', **font, color = 'white')
		plt.xticks(ind, tage_list, **font, color = 'white')
		plt.ylim(-round(max(avg_bar)*1.2,2), round(max(avg_bar)*1.2,2))
		plt.legend()
		ax.yaxis.grid(linestyle = 'dotted', color = 'white')

		plt.savefig(verzeichnis +'\ ' + bars_list[i] + '_Balken.png', dpi = 300, facecolor=fig.get_facecolor(), edgecolor='none')



dat = daten_einlesen('DAX30_1991-2018.xlsx')
#aus_wt_perf = Wochentage_Performance(dat[0], dat[1], dat[4])
#aus_wt_spa = Wochentage_Spanne(dat[0], dat[1], dat[2], dat[3])
#index_auf_wt_plot = index_wt_aufbereitung(aus_wt_perf)
#plot_wt_tage(index_auf_wt_plot, 'DAX')
#bar_Wt_tage(aus_wt_perf, 'DAX')
