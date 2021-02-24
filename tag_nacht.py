import openpyxl
import numpy as np
import os



def daten_einlesen(excel_datei):
	ws = excel_einlesen(excel_datei)
	rohdaten = rohdaten_einlesen_verarbeiten(ws)
	print('Daten eingelesen!')

	return rohdaten


def excel_einlesen(excel_datei):

	wb_daten = openpyxl.load_workbook(filename = excel_datei)
	ws_daten = wb_daten.active

	return [ws_daten]

# ---------------------------------------------------------------------------
# Retrun-Liste: [0] Datum; [1] Open; [2] High; [3] Low; [4] Close; [5] Differenz; [6] höchste Schwankung

def rohdaten_einlesen_verarbeiten(ws_rohdaten):

	daten_datum = spalte_einlesen(ws_rohdaten, 'A', 'Date')
	daten_open = spalte_einlesen(ws_rohdaten, 'B', 'Open')
	daten_high = spalte_einlesen(ws_rohdaten, 'C', 'High')
	daten_low = spalte_einlesen(ws_rohdaten, 'D', 'Low')
	daten_close = spalte_einlesen(ws_rohdaten, 'E', 'Close')

	daten_diff = []
	for i in range(len(daten_open)):
		daten_diff = daten_diff + [round((daten_close[i] - daten_open[i])/daten_open[i]*100,2)]

	daten_schw = []
	for i in range(len(daten_high)):
		daten_schw = daten_schw + [round((daten_high[i] - daten_low[i])/daten_open[i]*100,2)]



	return [daten_datum, daten_open, daten_high, daten_low, daten_close, daten_diff, daten_schw]


def spalte_einlesen(ws_rohdaten, spalte_name, name):
	ws = ws_rohdaten[0]
	spalten = ws[spalte_name]
	spalten_wert = []

	for i in range(len(spalten)):
		spalten_wert = spalten_wert + [spalten[i].value]

	spalten_wert.remove(name)

	return spalten_wert

#--------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------


def excel_ausgabe(daten_roh, daten_auswertung, basiswert):
	
	wb = workbook_erstellen()
	wb = daten_sheet(wb, daten_roh, daten_auswertung)
	workbook_speichern(wb, basiswert)

def workbook_erstellen():	
	wb = openpyxl.Workbook()

	return wb

def daten_sheet(wb, daten_roh, daten_aus):

	ws = wb.active
	ws.titel = 'Daten'

	da = daten_roh[0]
	op = daten_roh[1]
	low = daten_roh[2]
	high = daten_roh[3]
	clo = daten_roh[4]
	diff_tag = daten_aus[0]
	proz_tag = daten_aus[1]
	stetig_tag = daten_aus[2]
	index_tag = daten_aus[3]
	diff_nacht = daten_aus[4]
	proz_nacht = daten_aus[5]
	stetig_nacht = daten_aus[6]
	index_nacht = daten_aus[7]
	diff_gesamt = daten_aus[8]
	proz_gesamt = daten_aus[9]
	stetig_gesamt = daten_aus[10]
	index_gesamt = daten_aus[11]

	ws.cell(row = 1, column = 1, value = 'Date')
	ws.cell(row = 1, column = 2, value = 'Open')
	ws.cell(row = 1, column = 3, value = 'Low')
	ws.cell(row = 1, column = 4, value = 'High')
	ws.cell(row = 1, column = 5, value = 'Close')
	ws.cell(row = 1, column = 7, value = 'Differenz_Tag')
	ws.cell(row = 1, column = 8, value = 'Prozent_Tag')
	ws.cell(row = 1, column = 9, value = 'Stetig_Tag')
	ws.cell(row = 1, column = 10, value = 'Index_Tag')
	ws.cell(row = 1, column = 12, value = 'Differenz_Nacht')
	ws.cell(row = 1, column = 13, value = 'Prozent_Nacht')
	ws.cell(row = 1, column = 14, value = 'Stetig_Nacht')
	ws.cell(row = 1, column = 15, value = 'Index_Nacht')
	ws.cell(row = 1, column = 17, value = 'Differenz_Gesamt')
	ws.cell(row = 1, column = 18, value = 'Prozent_Gesamt')
	ws.cell(row = 1, column = 19, value = 'Stetig_Gesamt')
	ws.cell(row = 1, column = 20, value = 'Index_Gesamt')

	for i in range(len(daten_roh[0])):
		ws.cell(row = i+3, column = 1, value = da[i])
		ws.cell(row = i+3, column = 2, value = op[i])
		ws.cell(row = i+3, column = 3, value = low[i])
		ws.cell(row = i+3, column = 4, value = high[i])
		ws.cell(row = i+3, column = 5, value = clo[i])

	for i in range(len(daten_aus[0])):	
		ws.cell(row = i+3, column = 7, value = diff_tag[i])
		ws.cell(row = i+3, column = 8, value = proz_tag[i])
		ws.cell(row = i+3, column = 9, value = stetig_tag[i])

	for i in range(len(daten_aus[3])):
		ws.cell(row = i+2, column = 10, value = index_tag[i])

	for i in range(len(daten_aus[4])):	
		ws.cell(row = i+3, column = 12, value = diff_nacht[i])
		ws.cell(row = i+3, column = 13, value = proz_nacht[i])
		ws.cell(row = i+3, column = 14, value = stetig_nacht[i])

	for i in range(len(daten_aus[7])):
		ws.cell(row = i+2, column = 15, value = index_nacht[i])

	for i in range(len(daten_aus[8])):	
		ws.cell(row = i+3, column = 17, value = diff_gesamt[i])
		ws.cell(row = i+3, column = 18, value = proz_gesamt[i])
		ws.cell(row = i+3, column = 19, value = stetig_gesamt[i])

	for i in range(len(daten_aus[11])):
		ws.cell(row = i+2, column = 20, value = index_gesamt[i])


	ws.cell(row = 1, column = 22, value = 'Anzahl Tage')
	ws.cell(row = 2, column = 22, value = len(daten_aus[0]))
	ws.cell(row = 1, column = 23, value = 'Anzahl Nächte')
	ws.cell(row = 2, column = 23, value = len(daten_aus[4]))
	ws.cell(row = 1, column = 24, value = 'Anzahl Gesamt')
	ws.cell(row = 2, column = 24, value = len(daten_aus[8]))

	ws.cell(row = 4, column = 22, value = 'Punkte Tage')
	ws.cell(row = 5, column = 22, value = daten_aus[12])
	ws.cell(row = 4, column = 23, value = 'Punkte Nächte')
	ws.cell(row = 5, column = 23, value = daten_aus[13])
	ws.cell(row = 4, column = 24, value = 'Punkte Gesamt')
	ws.cell(row = 5, column = 24, value = daten_aus[14])
	ws.cell(row = 6, column = 24, value = 'Gesamt Punkte Delta')
	ws.cell(row = 7, column = 24, value = daten_aus[15])

	return wb

def workbook_speichern(w, basiswert):
	wb = w
	pfad1 = 'Output'
	pfad2 = basiswert + '_Tag_Nacht'
	verzeichnis = os.path.join(pfad1,pfad2)
	
	if not os.path.exists(verzeichnis):
		os.makedirs(verzeichnis)

	wb.save(verzeichnis +'\ ' + basiswert + '_Tag_Nacht_Verhältnis.xlsx')


#--------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------


def tag_nacht(ope, high, low, close):

	tag = []
	tag_stet = []
	tag_proz = []
	tag_index = [100]

	nacht = []
	nacht_stet =[]
	nacht_proz = []
	nacht_index = [100]

	gesamt = []
	gesamt_stet =[]
	gesamt_proz = []
	gesamt_index = [100]

	gesamt_letzter_tag = close[len(ope)-1] - ope[len(ope)-1]

	sum_diff_tag = 0
	sum_diff_nacht = 0
	sum_diff_gesamt = 0

	for i in range(len(ope)):

		diff = close[i] - ope[i]
		stet = (np.log(close[i]) - np.log(ope[i]))
		diff_proz = diff / ope[i]
		
		sum_diff_tag += diff
	
		tag = tag + [diff]
		tag_stet += [stet]
		tag_proz += [diff_proz]

	for i in range(len(ope)-1):

		diff = ope[i+1] - close[i]
		stet = (np.log(ope[i+1]) - np.log(close[i]))
		diff_proz = diff  / close[i]

		sum_diff_nacht += diff

		nacht = nacht + [diff]
		nacht_stet += [stet]
		nacht_proz += [diff_proz]

	for i in range(len(ope)-1):

		diff = ope[i+1] - ope[i]
		stet = (np.log(ope[i+1]) - np.log(ope[i]))
		diff_proz = diff / ope[i]

		sum_diff_gesamt += diff

		gesamt = gesamt + [diff]
		gesamt_stet += [stet]
		gesamt_proz += [diff_proz]

	for i in range(len(tag_stet)):

		ind = tag_index[i]*np.exp(tag_stet[i])
		
		tag_index += [ind]

	for i in range(len(nacht_stet)):

		ind = nacht_index[i]*np.exp(nacht_stet[i])
		
		nacht_index += [ind]

	for i in range(len(gesamt_stet)):

		ind = gesamt_index[i]*np.exp(gesamt_stet[i])
		
		gesamt_index += [ind]

	return [tag, tag_proz, tag_stet, tag_index, nacht, nacht_proz, nacht_stet, nacht_index, gesamt, gesamt_proz, gesamt_stet, gesamt_index, sum_diff_tag, sum_diff_nacht, sum_diff_gesamt,gesamt_letzter_tag]


dat = daten_einlesen('bitcoin.xlsx')
aus = tag_nacht(dat[1], dat[2], dat[3], dat[4])
excel_ausgabe(dat,aus,'bitcoin')