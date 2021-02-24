import openpyxl
import os

def excel_ausgabe(daten_roh, basiswert, quantile, lagepar, streupar, verh, pos_cl, neg_cl, pos_folg, neg_folg, kurslücken, strverh_pos, strverh_neg, strverh_span, wpd2, wa, ma):
	
	wb = workbook_erstellen()
	wb = daten_sheet(wb, daten_roh)
	wb = kennzahlen_quantile_sheet(wb, quantile)
	wb = kennzahlen_lagepar_sheet(wb, lagepar)
	wb = kennzahlen_streu_sheet(wb, streupar)
	wb = struktur_verhältnis_sheet(wb, verh, pos_cl, neg_cl)
	wb,r_a = struktur_folgen_sheet(wb, pos_folg, neg_folg)
	wb = struktur_kurslücken_sheet(wb, kurslücken[0], kurslücken[1], r_a)
	wb = strukturverhalten_sheet(wb, strverh_pos, strverh_neg, strverh_span)
	wb = wpd_sch_sheet(wb, wpd2)
	wb = kurslücken_sheet(wb, kurslücken[2], kurslücken[3], kurslücken[4], kurslücken[5], kurslücken[6])
	wb = wa_sheet(wb, wa)
	wb = ma_sheet(wb, ma)
	workbook_speichern(wb, basiswert)


def workbook_erstellen():	
	wb = openpyxl.Workbook()

	return wb


def daten_sheet(wb, daten_roh):

	ws = wb.active
	ws.titel = 'Daten'

	da = daten_roh[0]
	op = daten_roh[1]
	low = daten_roh[2]
	high = daten_roh[3]
	clo = daten_roh[4]
	dif = daten_roh[5]
	span = daten_roh[6]

	ws.cell(row = 1, column = 1, value = 'Date')
	ws.cell(row = 1, column = 2, value = 'Open')
	ws.cell(row = 1, column = 3, value = 'Low')
	ws.cell(row = 1, column = 4, value = 'High')
	ws.cell(row = 1, column = 5, value = 'Close')
	ws.cell(row = 1, column = 7, value = 'Differenz')
	ws.cell(row = 1, column = 8, value = 'Spanne')

	for i in range(len(daten_roh[0])):
		ws.cell(row = i+2, column = 1, value = da[i])
		ws.cell(row = i+2, column = 2, value = op[i])
		ws.cell(row = i+2, column = 3, value = low[i])
		ws.cell(row = i+2, column = 4, value = high[i])
		ws.cell(row = i+2, column = 5, value = clo[i])
		ws.cell(row = i+2, column = 7, value = dif[i])
		ws.cell(row = i+2, column = 8, value = span[i])

	return wb


def kennzahlen_quantile_sheet(wb, quantil):

	wb.create_sheet('Kennzahlen')
	ws = wb['Kennzahlen']

	ws.cell(row = 1, column = 1, value = 'Kennzahlen')
	ws.cell(row = 3, column = 1, value = 'Quantile:')
	ws.cell(row = 4, column = 1, value = 'Minimum')
	ws.cell(row = 5, column = 1, value = '25%')
	ws.cell(row = 6, column = 1, value = '50%')
	ws.cell(row = 7, column = 1, value = '75%')
	ws.cell(row = 8, column = 1, value = 'Maximum')
	ws.cell(row = 4, column = 2, value = quantil[0])
	ws.cell(row = 5, column = 2, value = quantil[1])
	ws.cell(row = 6, column = 2, value = quantil[2])
	ws.cell(row = 7, column = 2, value = quantil[3])
	ws.cell(row = 8, column = 2, value = quantil[4])

	return wb


def kennzahlen_lagepar_sheet(wb, lagepar):

	ws = wb['Kennzahlen']

	ws.cell(row = 10, column = 1, value = 'Lageparameter:')
	ws.cell(row = 11, column = 1, value = 'Modus')
	ws.cell(row = 12, column = 1, value = 'Median')
	ws.cell(row = 13, column = 1, value = 'arithmetisches Mittel')
	ws.cell(row = 11, column = 2, value = lagepar[0])
	ws.cell(row = 12, column = 2, value = lagepar[1])
	ws.cell(row = 13, column = 2, value = lagepar[2])

	return wb

def kennzahlen_streu_sheet(wb, streupar):

	ws = wb['Kennzahlen']

	ws.cell(row = 15, column = 1, value = 'Streuparameter:')
	ws.cell(row = 20, column = 1, value = 'Schiefe')
	ws.cell(row = 21, column = 1, value = 'Wölbung')
	ws.cell(row = 16, column = 1, value = 'Varianz')
	ws.cell(row = 17, column = 1, value = 'Volarität(StrAb')
	ws.cell(row = 18, column = 1, value = 'Semi-StrAb')
	ws.cell(row = 20, column = 2, value = streupar[0])
	ws.cell(row = 21, column = 2, value = streupar[1])
	ws.cell(row = 16, column = 2, value = streupar[2])
	ws.cell(row = 17, column = 2, value = streupar[3])
	ws.cell(row = 18, column = 2, value = streupar[4])

	var = streupar[5]
	ws.cell(row = 23, column = 1, value = 'VaR:')
	ws.cell(row = 24, column = 1, value = '95')
	ws.cell(row = 25, column = 1, value = '99')
	ws.cell(row = 26, column = 1, value = '99,9')
	ws.cell(row = 24, column = 2, value = var[0])
	ws.cell(row = 25, column = 2, value = var[1])
	ws.cell(row = 26, column = 2, value = var[2])   
	
	cvar = streupar[6]
	ws.cell(row = 28, column = 1, value = 'CVaR:')
	ws.cell(row = 29, column = 1, value = '95')
	ws.cell(row = 30, column = 1, value = '99')
	ws.cell(row = 31, column = 1, value = '99,9')
	ws.cell(row = 29, column = 2, value = cvar[0])
	ws.cell(row = 30, column = 2, value = cvar[1])
	ws.cell(row = 31, column = 2, value = cvar[2])

	return wb

def struktur_verhältnis_sheet(wb, verh, pos_cl, neg_cl):

	wb.create_sheet('Struktur')
	ws = wb['Struktur']

	ws.cell(row = 1, column = 1, value = 'Struktur')

	ws.cell(row = 3, column = 1, value = 'Verhältnis:')
	ws.cell(row = 4, column = 1, value = 'positive Tage')
	ws.cell(row = 5, column = 1, value = 'negative Tage')
	ws.cell(row = 6, column = 1, value = 'neutrale Tage')
	ws.cell(row = 7, column = 1, value = 'Total')
	ws.cell(row = 4, column = 2, value = verh[0])
	ws.cell(row = 5, column = 2, value = verh[1])
	ws.cell(row = 6, column = 2, value = verh[2])
	ws.cell(row = 7, column = 2, value = verh[0]+verh[1]+verh[2])
	ws.cell(row = 4, column = 3, value = round(verh[0]/(verh[0]+verh[1]+verh[2])*100,2))
	ws.cell(row = 5, column = 3, value = round(verh[1]/(verh[0]+verh[1]+verh[2])*100,2))
	ws.cell(row = 6, column = 3, value = round(verh[2]/(verh[0]+verh[1]+verh[2])*100,2))
	ws.cell(row = 7, column = 3, value = round((verh[0]+verh[1]+verh[2])/(verh[0]+verh[1]+verh[2])*100,2))

	ws.cell(row = 9, column = 1, value = 'positiver Performancebereich')
	ws.cell(row = 9, column = 2, value = 'absolute Häufigkeit')
	ws.cell(row = 9, column = 3, value = 'relative Häufigkeit')
	ws.cell(row = 10, column = 1, value = '0,01-0,50 %')
	ws.cell(row = 10, column = 2, value = pos_cl[0])
	ws.cell(row = 10, column = 3, value = round(pos_cl[0]/sum(pos_cl),4))
	ws.cell(row = 11, column = 1, value = '0,51-1,00 %')
	ws.cell(row = 11, column = 2, value = pos_cl[1])
	ws.cell(row = 11, column = 3, value = round(pos_cl[1]/sum(pos_cl),4))
	ws.cell(row = 12, column = 1, value = '1,01-1,50 %')
	ws.cell(row = 12, column = 2, value = pos_cl[2])
	ws.cell(row = 12, column = 3, value = round(pos_cl[2]/sum(pos_cl),4))
	ws.cell(row = 13, column = 1, value = '1,51-2,00 %')
	ws.cell(row = 13, column = 2, value = pos_cl[3])
	ws.cell(row = 13, column = 3, value = round(pos_cl[3]/sum(pos_cl),4))
	ws.cell(row = 14, column = 1, value = '>2,00 %')
	ws.cell(row = 14, column = 2, value = pos_cl[4])
	ws.cell(row = 14, column = 3, value = round(pos_cl[4]/sum(pos_cl),4))

	ws.cell(row = 16, column = 1, value = 'negativer Performancebereich')
	ws.cell(row = 16, column = 2, value = 'absolute Häufigkeit')
	ws.cell(row = 16, column = 3, value = 'relative Häufigkeit')
	ws.cell(row = 17, column = 1, value = '0,01-0,50 %')
	ws.cell(row = 17, column = 2, value = neg_cl[0])
	ws.cell(row = 17, column = 3, value = round(neg_cl[0]/sum(neg_cl),4))
	ws.cell(row = 18, column = 1, value = '0,51-1,00 %')
	ws.cell(row = 18, column = 2, value = neg_cl[1])
	ws.cell(row = 18, column = 3, value = round(neg_cl[1]/sum(neg_cl),4))
	ws.cell(row = 19, column = 1, value = '1,01-1,50 %')
	ws.cell(row = 19, column = 2, value = neg_cl[2])
	ws.cell(row = 19, column = 3, value = round(neg_cl[2]/sum(neg_cl),4))
	ws.cell(row = 20, column = 1, value = '1,51-2,00 %')
	ws.cell(row = 20, column = 2, value = neg_cl[3])
	ws.cell(row = 20, column = 3, value = round(neg_cl[3]/sum(neg_cl),4))
	ws.cell(row = 21, column = 1, value = '>2,00 %')
	ws.cell(row = 21, column = 2, value = neg_cl[4])
	ws.cell(row = 21, column = 3, value = round(neg_cl[4]/sum(neg_cl),4))

	abs_zahlen = sum(neg_cl) + sum(pos_cl) + verh[2]

	ws.cell(row = 9, column = 5, value = 'Webseite')
	ws.cell(row = 10, column = 5, value = 'Performancebereich')
	ws.cell(row = 10, column = 6, value = 'absolute Häufigkeit')
	ws.cell(row = 10, column = 7, value = 'relative Häufigkeit')
	ws.cell(row = 11, column = 5, value = 'kleiner -2,00 %')
	ws.cell(row = 11, column = 6, value = neg_cl[4])
	ws.cell(row = 11, column = 7, value = round(neg_cl[4]/abs_zahlen*100,2))
	ws.cell(row = 12, column = 5, value = 'zwischen -2,00 %' + ' und -1,51 %')
	ws.cell(row = 12, column = 6, value = neg_cl[3])
	ws.cell(row = 12, column = 7, value = round(neg_cl[3]/abs_zahlen*100,2))
	ws.cell(row = 13, column = 5, value = 'zwischen -1,50 %' + ' und -1,01 %')
	ws.cell(row = 13, column = 6, value = neg_cl[2])
	ws.cell(row = 13, column = 7, value = round(neg_cl[2]/abs_zahlen*100,2))
	ws.cell(row = 14, column = 5, value = 'zwischen -1,00 %' + ' und -0,51 %')
	ws.cell(row = 14, column = 6, value = neg_cl[1])
	ws.cell(row = 14, column = 7, value = round(neg_cl[1]/abs_zahlen*100,2))
	ws.cell(row = 15, column = 5, value = 'zwischen -0,50 %' + ' und -0,01 %')
	ws.cell(row = 15, column = 6, value = neg_cl[0])
	ws.cell(row = 15, column = 7, value = round(neg_cl[0]/abs_zahlen*100,2))
	ws.cell(row = 16, column = 5, value = '0,00 %')
	ws.cell(row = 16, column = 6, value = verh[2])
	ws.cell(row = 16, column = 7, value = round(verh[2]/abs_zahlen*100,2))
	ws.cell(row = 17, column = 5, value = 'zwischen 0,01 %' + ' und 0,50 %')
	ws.cell(row = 17, column = 6, value = pos_cl[0])
	ws.cell(row = 17, column = 7, value = round(pos_cl[0]/abs_zahlen*100,2))
	ws.cell(row = 18, column = 5, value = 'zwischen 0,51 %' + ' und 1,00 %')
	ws.cell(row = 18, column = 6, value = pos_cl[1])
	ws.cell(row = 18, column = 7, value = round(pos_cl[1]/abs_zahlen*100,2))
	ws.cell(row = 19, column = 5, value = 'zwischen 1,01 %' + ' und 1,50 %')
	ws.cell(row = 19, column = 6, value = pos_cl[2])
	ws.cell(row = 19, column = 7, value = round(pos_cl[2]/abs_zahlen*100,2))
	ws.cell(row = 20, column = 5, value = 'zwischen 1,51 %' + ' und 2,00 %')
	ws.cell(row = 20, column = 6, value = pos_cl[3])
	ws.cell(row = 20, column = 7, value = round(pos_cl[3]/abs_zahlen*100,2))
	ws.cell(row = 21, column = 5, value = 'größer 2,00 %')
	ws.cell(row = 21, column = 6, value = pos_cl[4])
	ws.cell(row = 21, column = 7, value = round(pos_cl[4]/abs_zahlen*100,2))

	return wb


def struktur_folgen_sheet(wb, pos_folgen, neg_folgen):

	ws = wb['Struktur']

	ws.cell(row = 23, column = 1, value = 'positive Folgen:')
	ws.cell(row = 24, column = 1, value = 'Folgentage')
	ws.cell(row = 25, column = 1, value = 'keine Folgentage')
	ws.cell(row = 26, column = 1, value = 'längste Folge')
	ws.cell(row = 24, column = 2, value = pos_folgen[1])
	ws.cell(row = 25, column = 2, value = pos_folgen[0]-pos_folgen[1])
	ws.cell(row = 26, column = 2, value = pos_folgen[2])
	ws.cell(row = 24, column = 3, value = round(pos_folgen[1]/(pos_folgen[0])*100,2))
	ws.cell(row = 25, column = 3, value = round((pos_folgen[0]-pos_folgen[1])/(pos_folgen[0])*100,2))
	
	ws.cell(row = 28, column = 1, value = 'Folgenlänge (in Tagen)')
	
	for i in range(1,8):
		if i == 7:
			ws.cell(row = 28+i, column = 1, value = '≥8')
		else:
			ws.cell(row =28+i , column = 1, value = i+1)

	ws.cell(row = 28, column = 2, value = 'Häufigkeit')

	pos_folg_häu = pos_folgen[3]
	for i in range(1,8):
		if i == 7:
			ws.cell(row = 28+i, column = 2, value = sum(pos_folg_häu[7:]))
		else:
			ws.cell(row = 28+i, column = 2, value = pos_folg_häu[i])

	rows_add = 8
	ws.cell(row = 29+rows_add, column = 1, value = 'negative Folgen:')
	ws.cell(row = 30+rows_add, column = 1, value = 'Folgentage')
	ws.cell(row = 31+rows_add, column = 1, value = 'keine Folgentage')
	ws.cell(row = 32+rows_add, column = 1, value = 'längste Folge')
	ws.cell(row = 30+rows_add, column = 2, value = neg_folgen[1])
	ws.cell(row = 31+rows_add, column = 2, value = neg_folgen[0]-neg_folgen[1])
	ws.cell(row = 32+rows_add, column = 2, value = neg_folgen[2])
	ws.cell(row = 30+rows_add, column = 3, value = round(neg_folgen[1]/(neg_folgen[0])*100,2))
	ws.cell(row = 31+rows_add, column = 3, value = round((neg_folgen[0]-neg_folgen[1])/(neg_folgen[0])*100,2))
	
	ws.cell(row = 34+rows_add, column = 1, value = 'Folgenlänge (in Tagen)')
	
	for i in range(1,8):
		if i == 7:
			ws.cell(row = 34+rows_add+i, column = 1, value = '≥8')
		else:
			ws.cell(row = 34+rows_add+i, column = 1, value = i+1)

	ws.cell(row = 34+rows_add, column = 2, value = 'Häufigkeit')

	neg_folg_häu = neg_folgen[3]
	for i in range(1,8):
		if i == 7:
			ws.cell(row = 34+rows_add+i, column = 2, value = sum(neg_folg_häu[7:]))
		else:
			ws.cell(row = 34+rows_add+i, column = 2, value = neg_folg_häu[i])

	rows_add = rows_add + 8

	return wb,rows_add


def struktur_kurslücken_sheet(wb, pos_kl, neg_kl, rows_add):

	ws = wb['Struktur']

	pos_kl_lä = pos_kl[3]
	pos_kl_anz = pos_kl[4]
	neg_kl_lä = neg_kl[3]
	neg_kl_anz = neg_kl[4]

	ws.cell(row = 35 + rows_add, column = 1, value = 'positive Kurslücken:')
	ws.cell(row = 36 + rows_add, column = 1, value = 'Kurslücken')
	ws.cell(row = 37 + rows_add, column = 1, value = 'bestättigte Kurslücken')
	ws.cell(row = 38 + rows_add, column = 1, value = 'offene Kurslücken')
	ws.cell(row = 39 + rows_add, column = 1, value = 'längste Kurslücke')
	ws.cell(row = 36 + rows_add, column = 2, value = pos_kl[0])
	ws.cell(row = 37 + rows_add, column = 2, value = pos_kl[1])
	ws.cell(row = 38 + rows_add, column = 2, value = pos_kl[2])
	ws.cell(row = 39 + rows_add, column = 2, value = pos_kl_lä[len(pos_kl_lä)-1])
	
	ws.cell(row = 41 + rows_add, column = 1, value = 'Kurslückenlänge')

	for i in range(11):
		if i == 10:
			ws.cell(row = 42+rows_add+i, column = 1, value = '≥10')
		else:
			ws.cell(row = 42+rows_add+i, column = 1, value = pos_kl_lä[i])


	ws.cell(row = 41+rows_add, column = 2, value = 'Kurslückenhäufigkeit')

	for i in range(11):
		if i == 10:
			ws.cell(row = 42+rows_add+i, column = 2, value = sum(pos_kl_anz[10:]))
		else:
			ws.cell(row = 42+rows_add+i, column = 2, value = pos_kl_anz[i])

	rows_add = rows_add + 11

	ws.cell(row = 44 + rows_add, column = 1, value = 'negative Kurslücken:')
	ws.cell(row = 45 + rows_add, column = 1, value = 'Kurslücken')
	ws.cell(row = 46 + rows_add, column = 1, value = 'bestättigte Kurslücken')
	ws.cell(row = 47 + rows_add, column = 1, value = 'offene Kurslücken')
	ws.cell(row = 48 + rows_add, column = 1, value = 'längste Kurslücken')
	ws.cell(row = 45 + rows_add, column = 2, value = neg_kl[0])
	ws.cell(row = 46 + rows_add, column = 2, value = neg_kl[1])
	ws.cell(row = 47 + rows_add, column = 2, value = neg_kl[2])
	ws.cell(row = 48 + rows_add, column = 2, value = neg_kl_lä[len(neg_kl_lä)-1])
	
	ws.cell(row = 50 + rows_add, column = 1, value = 'Kurslückenlänge')

	for i in range(11):
		if i == 10:
			ws.cell(row = 51+rows_add+i, column = 1, value = '≥10')
		else:
			ws.cell(row = 51+rows_add+i, column = 1, value = neg_kl_lä[i])

	ws.cell(row = 50 + rows_add, column = 2, value = 'Kurslückenhäufigkeit')

	for i in range(11):
		if i == 10:
			ws.cell(row = 51+rows_add+i, column = 2, value = sum(neg_kl_anz[10:]))
		else:
			ws.cell(row = 51+rows_add+i, column = 2, value = neg_kl_anz[i])

	return wb

def strukturverhalten_sheet(wb, strverh_pos, strverh_neg, strverh_span):

	wb.create_sheet('Strukturverhalten')
	ws = wb['Strukturverhalten']

	ws.cell(row = 1, column = 1, value = 'Strukturverhalten')
	ws.cell(row = 3, column = 1, value = 'positive Tage:')
	ws.cell(row = 4, column = 1, value = 'Minimum')
	ws.cell(row = 5, column = 1, value = 'Durchschnitt')
	ws.cell(row = 6, column = 1, value = 'Maximum')
	ws.cell(row = 4, column = 2, value = strverh_pos[0])
	ws.cell(row = 5, column = 2, value = strverh_pos[1])
	ws.cell(row = 6, column = 2, value = strverh_pos[2])

	ws.cell(row = 8, column = 1, value = 'negative Tage:')
	ws.cell(row = 9, column = 1, value = 'Minimum')
	ws.cell(row = 10, column = 1, value = 'Durchschnitt')
	ws.cell(row = 11, column = 1, value = 'Maximum')
	ws.cell(row = 9, column = 2, value = strverh_neg[0])
	ws.cell(row = 10, column = 2, value = strverh_neg[1])
	ws.cell(row = 11, column = 2, value = strverh_neg[2])

	ws.cell(row = 13, column = 1, value = 'Handelsspanne:')
	ws.cell(row = 14, column = 1, value = 'Minimum')
	ws.cell(row = 15, column = 1, value = 'Durchschnitt')
	ws.cell(row = 16, column = 1, value = 'Maximum')
	ws.cell(row = 14, column = 2, value = strverh_span[0])
	ws.cell(row = 15, column = 2, value = strverh_span[1])
	ws.cell(row = 16, column = 2, value = strverh_span[2])

	return wb


def wpd_sheet(wb, wpd):

	wb.create_sheet('Was passiert')
	ws = wb['Was passiert']

	ws.cell(row = 1, column = 1, value = 'Was passiert')

	ws.cell(row = 3, column = 1, value = 'Bereich')
	ws.cell(row = 3, column = 2, value = 'absolute Häufigkeit')
	ws.cell(row = 3, column = 3, value = 'relative Häufigkeit')
	ws.cell(row = 4, column = 1, value = '0,01-0,50 %')
	ws.cell(row = 4, column = 2, value = wpd[0][0])
	ws.cell(row = 5, column = 1, value = '0,51-1,00 %')
	ws.cell(row = 5, column = 2, value = wpd[0][1])
	ws.cell(row = 6, column = 1, value = '1,01-1,50 %')
	ws.cell(row = 6, column = 2, value = wpd[0][2])
	ws.cell(row = 7, column = 1, value = '1,51-2,00 %')
	ws.cell(row = 7, column = 2, value = wpd[0][3])
	ws.cell(row = 8, column = 1, value = '>2,00 %')
	ws.cell(row = 8, column = 2, value = wpd[0][4])

	ws.cell(row = 10, column = 1, value = '0,01-0,50 %')
	ws.cell(row = 11, column = 1, value = 'ein Tag (intraday)')
	ws.cell(row = 12, column = 1, value = 'Bereich')
	ws.cell(row = 12, column = 2, value = 'Häufigkeit')
	ws.cell(row = 13, column = 1, value = '< -2')
	ws.cell(row = 13, column = 2, value = wpd[1][0][0])
	ws.cell(row = 14, column = 1, value = '-2,00 - -1,51 %')
	ws.cell(row = 14, column = 2, value = wpd[1][0][1])
	ws.cell(row = 15, column = 1, value = '-1,50 - -1,01 %')
	ws.cell(row = 15, column = 2, value = wpd[1][0][2])
	ws.cell(row = 16, column = 1, value = '-1,00 - -0,51 %')
	ws.cell(row = 16, column = 2, value = wpd[1][0][3])
	ws.cell(row = 17, column = 1, value = '-0,50 - -0,01 %')
	ws.cell(row = 17, column = 2, value = wpd[1][0][4])
	ws.cell(row = 18, column = 1, value = '0 %')
	ws.cell(row = 18, column = 2, value = wpd[1][0][5])
	ws.cell(row = 19, column = 1, value = '0,01 - 0,50 %')
	ws.cell(row = 19, column = 2, value = wpd[1][0][6])
	ws.cell(row = 20, column = 1, value = '0,51 - 1,00 %')
	ws.cell(row = 20, column = 2, value = wpd[1][0][7])
	ws.cell(row = 21, column = 1, value = '1,01 - 1,50 %')
	ws.cell(row = 21, column = 2, value = wpd[1][0][8])
	ws.cell(row = 22, column = 1, value = '1,51 - 2,00 %')
	ws.cell(row = 22, column = 2, value = wpd[1][0][9])
	ws.cell(row = 23, column = 1, value = '>2 %')
	ws.cell(row = 23, column = 2, value = wpd[1][0][10])

	ws.cell(row = 25, column = 1, value = 'fünf Tage (intraday)')
	ws.cell(row = 26, column = 1, value = 'Bereich')
	ws.cell(row = 26, column = 2, value = 'Häufigkeit')
	ws.cell(row = 27, column = 1, value = '< -2')
	ws.cell(row = 27, column = 2, value = wpd[2][0][0])
	ws.cell(row = 28, column = 1, value = '-2,00 - -1,51 %')
	ws.cell(row = 28, column = 2, value = wpd[2][0][1])
	ws.cell(row = 29, column = 1, value = '-1,50 - -1,01 %')
	ws.cell(row = 29, column = 2, value = wpd[2][0][2])
	ws.cell(row = 30, column = 1, value = '-1,00 - -0,51 %')
	ws.cell(row = 30, column = 2, value = wpd[2][0][3])
	ws.cell(row = 31, column = 1, value = '-0,50 - -0,01 %')
	ws.cell(row = 31, column = 2, value = wpd[2][0][4])
	ws.cell(row = 32, column = 1, value = '0 %')
	ws.cell(row = 32, column = 2, value = wpd[2][0][5])
	ws.cell(row = 33, column = 1, value = '0,01 - 0,50 %')
	ws.cell(row = 33, column = 2, value = wpd[2][0][6])
	ws.cell(row = 34, column = 1, value = '0,51 - 1,00 %')
	ws.cell(row = 34, column = 2, value = wpd[2][0][7])
	ws.cell(row = 35, column = 1, value = '1,01 - 1,50 %')
	ws.cell(row = 35, column = 2, value = wpd[2][0][8])
	ws.cell(row = 36, column = 1, value = '1,51 - 2,00 %')
	ws.cell(row = 36, column = 2, value = wpd[2][0][9])
	ws.cell(row = 37, column = 1, value = '>2 %')
	ws.cell(row = 37, column = 2, value = wpd[2][0][10])

	ws.cell(row = 39, column = 1, value = 'dreißig Tage (intraday)')
	ws.cell(row = 40, column = 1, value = 'Bereich')
	ws.cell(row = 40, column = 2, value = 'Häufigkeit')
	ws.cell(row = 41, column = 1, value = '< -2')
	ws.cell(row = 41, column = 2, value = wpd[3][0][0])
	ws.cell(row = 42, column = 1, value = '-2,00 - -1,51 %')
	ws.cell(row = 42, column = 2, value = wpd[3][0][1])
	ws.cell(row = 43, column = 1, value = '-1,50 - -1,01 %')
	ws.cell(row = 43, column = 2, value = wpd[3][0][2])
	ws.cell(row = 44, column = 1, value = '-1,00 - -0,51 %')
	ws.cell(row = 44, column = 2, value = wpd[3][0][3])
	ws.cell(row = 45, column = 1, value = '-0,50 - -0,01 %')
	ws.cell(row = 45, column = 2, value = wpd[3][0][4])
	ws.cell(row = 46, column = 1, value = '0 %')
	ws.cell(row = 46, column = 2, value = wpd[3][0][5])
	ws.cell(row = 47, column = 1, value = '0,01 - 0,50 %')
	ws.cell(row = 47, column = 2, value = wpd[3][0][6])
	ws.cell(row = 48, column = 1, value = '0,51 - 1,00 %')
	ws.cell(row = 48, column = 2, value = wpd[3][0][7])
	ws.cell(row = 49, column = 1, value = '1,01 - 1,50 %')
	ws.cell(row = 49, column = 2, value = wpd[3][0][8])
	ws.cell(row = 50, column = 1, value = '1,51 - 2,00 %')
	ws.cell(row = 50, column = 2, value = wpd[3][0][9])
	ws.cell(row = 51, column = 1, value = '>2 %')
	ws.cell(row = 51, column = 2, value = wpd[3][0][10])

	ws.cell(row = 26, column = 1, value = '0,51 - 1,00 %')
	ws.cell(row = 27, column = 1, value = 'ein Tag (intraday)')
	ws.cell(row = 28, column = 1, value = 'Bereich')
	ws.cell(row = 28, column = 2, value = 'Häufigkeit')
	ws.cell(row = 29, column = 1, value = '< -2')
	ws.cell(row = 29, column = 2, value = wpd[1][1][0])
	ws.cell(row = 30, column = 1, value = '-2,00 - -1,51 %')
	ws.cell(row = 30, column = 2, value = wpd[1][1][1])
	ws.cell(row = 31, column = 1, value = '-1,50 - -1,01 %')
	ws.cell(row = 31, column = 2, value = wpd[1][1][2])
	ws.cell(row = 32, column = 1, value = '-1,00 - -0,51 %')
	ws.cell(row = 32, column = 2, value = wpd[1][1][3])
	ws.cell(row = 33, column = 1, value = '-0,50 - -0,01 %')
	ws.cell(row = 33, column = 2, value = wpd[1][1][4])
	ws.cell(row = 34, column = 1, value = '0 %')
	ws.cell(row = 34, column = 2, value = wpd[1][1][5])
	ws.cell(row = 35, column = 1, value = '0,01 - 0,50 %')
	ws.cell(row = 35, column = 2, value = wpd[1][1][6])
	ws.cell(row = 36, column = 1, value = '0,51 - 1,00 %')
	ws.cell(row = 36, column = 2, value = wpd[1][1][7])
	ws.cell(row = 37, column = 1, value = '1,01 - 1,50 %')
	ws.cell(row = 37, column = 2, value = wpd[1][1][8])
	ws.cell(row = 38, column = 1, value = '1,51 - 2,00 %')
	ws.cell(row = 38, column = 2, value = wpd[1][1][9])
	ws.cell(row = 39, column = 1, value = '>2 %')
	ws.cell(row = 39, column = 2, value = wpd[1][1][10])

	ws.cell(row = 27, column = 4, value = 'fünf Tage (intraday)')
	ws.cell(row = 28, column = 4, value = 'Bereich')
	ws.cell(row = 28, column = 5, value = 'Häufigkeit')
	ws.cell(row = 29, column = 4, value = '< -2')
	ws.cell(row = 29, column = 5, value = wpd[2][1][0])
	ws.cell(row = 30, column = 4, value = '-2,00 - -1,51 %')
	ws.cell(row = 30, column = 5, value = wpd[2][1][1])
	ws.cell(row = 31, column = 4, value = '-1,50 - -1,01 %')
	ws.cell(row = 31, column = 5, value = wpd[2][1][2])
	ws.cell(row = 32, column = 4, value = '-1,00 - -0,51 %')
	ws.cell(row = 32, column = 5, value = wpd[2][1][3])
	ws.cell(row = 33, column = 4, value = '-0,50 - -0,01 %')
	ws.cell(row = 33, column = 5, value = wpd[2][1][4])
	ws.cell(row = 34, column = 4, value = '0 %')
	ws.cell(row = 34, column = 5, value = wpd[2][1][5])
	ws.cell(row = 35, column = 4, value = '0,01 - 0,50 %')
	ws.cell(row = 35, column = 5, value = wpd[2][1][6])
	ws.cell(row = 36, column = 4, value = '0,51 - 1,00 %')
	ws.cell(row = 36, column = 5, value = wpd[2][1][7])
	ws.cell(row = 37, column = 4, value = '1,01 - 1,50 %')
	ws.cell(row = 37, column = 5, value = wpd[2][1][8])
	ws.cell(row = 38, column = 4, value = '1,51 - 2,00 %')
	ws.cell(row = 38, column = 5, value = wpd[2][1][9])
	ws.cell(row = 39, column = 4, value = '>2 %')
	ws.cell(row = 39, column = 5, value = wpd[2][1][10])

	ws.cell(row = 27, column = 7, value = 'dreißig Tage (intraday)')
	ws.cell(row = 28, column = 7, value = 'Bereich')
	ws.cell(row = 28, column = 8, value = 'Häufigkeit')
	ws.cell(row = 29, column = 7, value = '< -2')
	ws.cell(row = 29, column = 8, value = wpd[3][1][0])
	ws.cell(row = 30, column = 7, value = '-2,00 - -1,51 %')
	ws.cell(row = 30, column = 8, value = wpd[3][1][1])
	ws.cell(row = 31, column = 7, value = '-1,50 - -1,01 %')
	ws.cell(row = 31, column = 8, value = wpd[3][1][2])
	ws.cell(row = 32, column = 7, value = '-1,00 - -0,51 %')
	ws.cell(row = 32, column = 8, value = wpd[3][1][3])
	ws.cell(row = 33, column = 7, value = '-0,50 - -0,01 %')
	ws.cell(row = 33, column = 8, value = wpd[3][1][4])
	ws.cell(row = 34, column = 7, value = '0 %')
	ws.cell(row = 34, column = 8, value = wpd[3][1][5])
	ws.cell(row = 35, column = 7, value = '0,01 - 0,50 %')
	ws.cell(row = 35, column = 8, value = wpd[3][1][6])
	ws.cell(row = 36, column = 7, value = '0,51 - 1,00 %')
	ws.cell(row = 36, column = 8, value = wpd[3][1][7])
	ws.cell(row = 37, column = 7, value = '1,01 - 1,50 %')
	ws.cell(row = 37, column = 8, value = wpd[3][1][8])
	ws.cell(row = 38, column = 7, value = '1,51 - 2,00 %')
	ws.cell(row = 38, column = 8, value = wpd[3][1][9])
	ws.cell(row = 39, column = 7, value = '>2 %')
	ws.cell(row = 39, column = 8, value = wpd[3][1][10])

	ws.cell(row = 42, column = 1, value = '1,01-1,50 %')
	ws.cell(row = 43, column = 1, value = 'ein Tag (intraday)')
	ws.cell(row = 44, column = 1, value = 'Bereich')
	ws.cell(row = 44, column = 2, value = 'Häufigkeit')
	ws.cell(row = 45, column = 1, value = '< -2')
	ws.cell(row = 45, column = 2, value = wpd[1][2][0])
	ws.cell(row = 46, column = 1, value = '-2,00 - -1,51 %')
	ws.cell(row = 46, column = 2, value = wpd[1][2][1])
	ws.cell(row = 47, column = 1, value = '-1,50 - -1,01 %')
	ws.cell(row = 47, column = 2, value = wpd[1][2][2])
	ws.cell(row = 48, column = 1, value = '-1,00 - -0,51 %')
	ws.cell(row = 48, column = 2, value = wpd[1][2][3])
	ws.cell(row = 49, column = 1, value = '-0,50 - -0,01 %')
	ws.cell(row = 49, column = 2, value = wpd[1][2][4])
	ws.cell(row = 50, column = 1, value = '0 %')
	ws.cell(row = 50, column = 2, value = wpd[1][2][5])
	ws.cell(row = 51, column = 1, value = '0,01 - 0,50 %')
	ws.cell(row = 51, column = 2, value = wpd[1][2][6])
	ws.cell(row = 52, column = 1, value = '0,51 - 1,00 %')
	ws.cell(row = 52, column = 2, value = wpd[1][2][7])
	ws.cell(row = 53, column = 1, value = '1,01 - 1,50 %')
	ws.cell(row = 53, column = 2, value = wpd[1][2][8])
	ws.cell(row = 54, column = 1, value = '1,51 - 2,00 %')
	ws.cell(row = 54, column = 2, value = wpd[1][2][9])
	ws.cell(row = 55, column = 1, value = '>2 %')
	ws.cell(row = 55, column = 2, value = wpd[1][2][10])

	ws.cell(row = 43, column = 4, value = 'fünf Tage (intraday)')
	ws.cell(row = 44, column = 4, value = 'Bereich')
	ws.cell(row = 44, column = 5, value = 'Häufigkeit')
	ws.cell(row = 45, column = 4, value = '< -2')
	ws.cell(row = 45, column = 5, value = wpd[2][2][0])
	ws.cell(row = 46, column = 4, value = '-2,00 - -1,51 %')
	ws.cell(row = 46, column = 5, value = wpd[2][2][1])
	ws.cell(row = 47, column = 4, value = '-1,50 - -1,01 %')
	ws.cell(row = 47, column = 5, value = wpd[2][2][2])
	ws.cell(row = 48, column = 4, value = '-1,00 - -0,51 %')
	ws.cell(row = 48, column = 5, value = wpd[2][2][3])
	ws.cell(row = 49, column = 4, value = '-0,50 - -0,01 %')
	ws.cell(row = 49, column = 5, value = wpd[2][2][4])
	ws.cell(row = 50, column = 4, value = '0 %')
	ws.cell(row = 50, column = 5, value = wpd[2][2][5])
	ws.cell(row = 51, column = 4, value = '0,01 - 0,50 %')
	ws.cell(row = 51, column = 5, value = wpd[2][2][6])
	ws.cell(row = 52, column = 4, value = '0,51 - 1,00 %')
	ws.cell(row = 52, column = 5, value = wpd[2][2][7])
	ws.cell(row = 53, column = 4, value = '1,01 - 1,50 %')
	ws.cell(row = 53, column = 5, value = wpd[2][2][8])
	ws.cell(row = 54, column = 4, value = '1,51 - 2,00 %')
	ws.cell(row = 54, column = 5, value = wpd[2][2][9])
	ws.cell(row = 55, column = 4, value = '>2 %')
	ws.cell(row = 55, column = 5, value = wpd[2][2][10])

	ws.cell(row = 43, column = 7, value = 'dreißig Tage (intraday)')
	ws.cell(row = 44, column = 7, value = 'Bereich')
	ws.cell(row = 44, column = 8, value = 'Häufigkeit')
	ws.cell(row = 45, column = 7, value = '< -2')
	ws.cell(row = 45, column = 8, value = wpd[3][2][0])
	ws.cell(row = 46, column = 7, value = '-2,00 - -1,51 %')
	ws.cell(row = 46, column = 8, value = wpd[3][2][1])
	ws.cell(row = 47, column = 7, value = '-1,50 - -1,01 %')
	ws.cell(row = 47, column = 8, value = wpd[3][2][2])
	ws.cell(row = 48, column = 7, value = '-1,00 - -0,51 %')
	ws.cell(row = 48, column = 8, value = wpd[3][2][3])
	ws.cell(row = 49, column = 7, value = '-0,50 - -0,01 %')
	ws.cell(row = 49, column = 8, value = wpd[3][2][4])
	ws.cell(row = 50, column = 7, value = '0 %')
	ws.cell(row = 50, column = 8, value = wpd[3][2][5])
	ws.cell(row = 51, column = 7, value = '0,01 - 0,50 %')
	ws.cell(row = 51, column = 8, value = wpd[3][2][6])
	ws.cell(row = 52, column = 7, value = '0,51 - 1,00 %')
	ws.cell(row = 52, column = 8, value = wpd[3][2][7])
	ws.cell(row = 53, column = 7, value = '1,01 - 1,50 %')
	ws.cell(row = 53, column = 8, value = wpd[3][2][8])
	ws.cell(row = 54, column = 7, value = '1,51 - 2,00 %')
	ws.cell(row = 54, column = 8, value = wpd[3][2][9])
	ws.cell(row = 55, column = 7, value = '>2 %')
	ws.cell(row = 55, column = 8, value = wpd[3][2][10])

	ws.cell(row = 58, column = 1, value = '1,51-2,00 %')
	ws.cell(row = 59, column = 1, value = 'ein Tag (intraday)')
	ws.cell(row = 60, column = 1, value = 'Bereich')
	ws.cell(row = 60, column = 2, value = 'Häufigkeit')
	ws.cell(row = 61, column = 1, value = '< -2')
	ws.cell(row = 61, column = 2, value = wpd[1][3][0])
	ws.cell(row = 62, column = 1, value = '-2,00 - -1,51 %')
	ws.cell(row = 62, column = 2, value = wpd[1][3][1])
	ws.cell(row = 63, column = 1, value = '-1,50 - -1,01 %')
	ws.cell(row = 63, column = 2, value = wpd[1][3][2])
	ws.cell(row = 64, column = 1, value = '-1,00 - -0,51 %')
	ws.cell(row = 64, column = 2, value = wpd[1][3][3])
	ws.cell(row = 65, column = 1, value = '-0,50 - -0,01 %')
	ws.cell(row = 65, column = 2, value = wpd[1][3][4])
	ws.cell(row = 66, column = 1, value = '0 %')
	ws.cell(row = 66, column = 2, value = wpd[1][3][5])
	ws.cell(row = 67, column = 1, value = '0,01 - 0,50 %')
	ws.cell(row = 67, column = 2, value = wpd[1][3][6])
	ws.cell(row = 68, column = 1, value = '0,51 - 1,00 %')
	ws.cell(row = 68, column = 2, value = wpd[1][3][7])
	ws.cell(row = 69, column = 1, value = '1,01 - 1,50 %')
	ws.cell(row = 69, column = 2, value = wpd[1][3][8])
	ws.cell(row = 70, column = 1, value = '1,51 - 2,00 %')
	ws.cell(row = 70, column = 2, value = wpd[1][3][9])
	ws.cell(row = 71, column = 1, value = '>2 %')
	ws.cell(row = 71, column = 2, value = wpd[1][3][10])

	ws.cell(row = 59, column = 4, value = 'fünf Tage (intraday)')
	ws.cell(row = 60, column = 4, value = 'Bereich')
	ws.cell(row = 60, column = 5, value = 'Häufigkeit')
	ws.cell(row = 61, column = 4, value = '< -2')
	ws.cell(row = 61, column = 5, value = wpd[2][3][0])
	ws.cell(row = 62, column = 4, value = '-2,00 - -1,51 %')
	ws.cell(row = 62, column = 5, value = wpd[2][3][1])
	ws.cell(row = 63, column = 4, value = '-1,50 - -1,01 %')
	ws.cell(row = 63, column = 5, value = wpd[2][3][2])
	ws.cell(row = 64, column = 4, value = '-1,00 - -0,51 %')
	ws.cell(row = 64, column = 5, value = wpd[2][3][3])
	ws.cell(row = 65, column = 4, value = '-0,50 - -0,01 %')
	ws.cell(row = 65, column = 5, value = wpd[2][3][4])
	ws.cell(row = 66, column = 4, value = '0 %')
	ws.cell(row = 66, column = 5, value = wpd[2][3][5])
	ws.cell(row = 67, column = 4, value = '0,01 - 0,50 %')
	ws.cell(row = 67, column = 5, value = wpd[2][3][6])
	ws.cell(row = 68, column = 4, value = '0,51 - 1,00 %')
	ws.cell(row = 68, column = 5, value = wpd[2][3][7])
	ws.cell(row = 69, column = 4, value = '1,01 - 1,50 %')
	ws.cell(row = 69, column = 5, value = wpd[2][3][8])
	ws.cell(row = 70, column = 4, value = '1,51 - 2,00 %')
	ws.cell(row = 70, column = 5, value = wpd[2][3][9])
	ws.cell(row = 71, column = 4, value = '>2 %')
	ws.cell(row = 71, column = 5, value = wpd[2][3][10])

	ws.cell(row = 59, column = 7, value = 'dreißig Tage (intraday)')
	ws.cell(row = 60, column = 7, value = 'Bereich')
	ws.cell(row = 60, column = 8, value = 'Häufigkeit')
	ws.cell(row = 61, column = 7, value = '< -2')
	ws.cell(row = 61, column = 8, value = wpd[3][3][0])
	ws.cell(row = 62, column = 7, value = '-2,00 - -1,51 %')
	ws.cell(row = 62, column = 8, value = wpd[3][3][1])
	ws.cell(row = 63, column = 7, value = '-1,50 - -1,01 %')
	ws.cell(row = 63, column = 8, value = wpd[3][3][2])
	ws.cell(row = 64, column = 7, value = '-1,00 - -0,51 %')
	ws.cell(row = 64, column = 8, value = wpd[3][3][3])
	ws.cell(row = 65, column = 7, value = '-0,50 - -0,01 %')
	ws.cell(row = 65, column = 8, value = wpd[3][3][4])
	ws.cell(row = 66, column = 7, value = '0 %')
	ws.cell(row = 66, column = 8, value = wpd[3][3][5])
	ws.cell(row = 67, column = 7, value = '0,01 - 0,50 %')
	ws.cell(row = 67, column = 8, value = wpd[3][3][6])
	ws.cell(row = 68, column = 7, value = '0,51 - 1,00 %')
	ws.cell(row = 68, column = 8, value = wpd[3][3][7])
	ws.cell(row = 69, column = 7, value = '1,01 - 1,50 %')
	ws.cell(row = 69, column = 8, value = wpd[3][3][8])
	ws.cell(row = 70, column = 7, value = '1,51 - 2,00 %')
	ws.cell(row = 70, column = 8, value = wpd[3][3][9])
	ws.cell(row = 71, column = 7, value = '>2 %')
	ws.cell(row = 71, column = 8, value = wpd[3][3][10])

	ws.cell(row = 74, column = 1, value = '>2,00 %')
	ws.cell(row = 75, column = 1, value = 'ein Tag (intraday)')
	ws.cell(row = 76, column = 1, value = 'Bereich')
	ws.cell(row = 76, column = 2, value = 'Häufigkeit')
	ws.cell(row = 77, column = 1, value = '< -2')
	ws.cell(row = 77, column = 2, value = wpd[1][4][0])
	ws.cell(row = 78, column = 1, value = '-2,00 - -1,51 %')
	ws.cell(row = 78, column = 2, value = wpd[1][4][1])
	ws.cell(row = 79, column = 1, value = '-1,50 - -1,01 %')
	ws.cell(row = 79, column = 2, value = wpd[1][4][2])
	ws.cell(row = 80, column = 1, value = '-1,00 - -0,51 %')
	ws.cell(row = 80, column = 2, value = wpd[1][4][3])
	ws.cell(row = 81, column = 1, value = '-0,50 - -0,01 %')
	ws.cell(row = 81, column = 2, value = wpd[1][4][4])
	ws.cell(row = 82, column = 1, value = '0 %')
	ws.cell(row = 82, column = 2, value = wpd[1][4][5])
	ws.cell(row = 83, column = 1, value = '0,01 - 0,50 %')
	ws.cell(row = 83, column = 2, value = wpd[1][4][6])
	ws.cell(row = 84, column = 1, value = '0,51 - 1,00 %')
	ws.cell(row = 84, column = 2, value = wpd[1][4][7])
	ws.cell(row = 85, column = 1, value = '1,01 - 1,50 %')
	ws.cell(row = 85, column = 2, value = wpd[1][4][8])
	ws.cell(row = 86, column = 1, value = '1,51 - 2,00 %')
	ws.cell(row = 86, column = 2, value = wpd[1][4][9])
	ws.cell(row = 87, column = 1, value = '>2 %')
	ws.cell(row = 87, column = 2, value = wpd[1][4][10])

	ws.cell(row = 75, column = 4, value = 'fünf Tage (intraday)')
	ws.cell(row = 76, column = 4, value = 'Bereich')
	ws.cell(row = 76, column = 5, value = 'Häufigkeit')
	ws.cell(row = 77, column = 4, value = '< -2')
	ws.cell(row = 77, column = 5, value = wpd[2][4][0])
	ws.cell(row = 78, column = 4, value = '-2,00 - -1,51 %')
	ws.cell(row = 78, column = 5, value = wpd[2][4][1])
	ws.cell(row = 79, column = 4, value = '-1,50 - -1,01 %')
	ws.cell(row = 79, column = 5, value = wpd[2][4][2])
	ws.cell(row = 80, column = 4, value = '-1,00 - -0,51 %')
	ws.cell(row = 80, column = 5, value = wpd[2][4][3])
	ws.cell(row = 81, column = 4, value = '-0,50 - -0,01 %')
	ws.cell(row = 81, column = 5, value = wpd[2][4][4])
	ws.cell(row = 82, column = 4, value = '0 %')
	ws.cell(row = 82, column = 5, value = wpd[2][4][5])
	ws.cell(row = 83, column = 4, value = '0,01 - 0,50 %')
	ws.cell(row = 83, column = 5, value = wpd[2][4][6])
	ws.cell(row = 84, column = 4, value = '0,51 - 1,00 %')
	ws.cell(row = 84, column = 5, value = wpd[2][4][7])
	ws.cell(row = 85, column = 4, value = '1,01 - 1,50 %')
	ws.cell(row = 85, column = 5, value = wpd[2][4][8])
	ws.cell(row = 86, column = 4, value = '1,51 - 2,00 %')
	ws.cell(row = 86, column = 5, value = wpd[2][4][9])
	ws.cell(row = 87, column = 4, value = '>2 %')
	ws.cell(row = 87, column = 5, value = wpd[2][4][10])

	ws.cell(row = 75, column = 7, value = 'dreißig Tage (intraday)')
	ws.cell(row = 76, column = 7, value = 'Bereich')
	ws.cell(row = 76, column = 8, value = 'Häufigkeit')
	ws.cell(row = 77, column = 7, value = '< -2')
	ws.cell(row = 77, column = 8, value = wpd[3][4][0])
	ws.cell(row = 78, column = 7, value = '-2,00 - -1,51 %')
	ws.cell(row = 78, column = 8, value = wpd[3][4][1])
	ws.cell(row = 79, column = 7, value = '-1,50 - -1,01 %')
	ws.cell(row = 79, column = 8, value = wpd[3][4][2])
	ws.cell(row = 80, column = 7, value = '-1,00 - -0,51 %')
	ws.cell(row = 80, column = 8, value = wpd[3][4][3])
	ws.cell(row = 81, column = 7, value = '-0,50 - -0,01 %')
	ws.cell(row = 81, column = 8, value = wpd[3][4][4])
	ws.cell(row = 82, column = 7, value = '0 %')
	ws.cell(row = 82, column = 8, value = wpd[3][4][5])
	ws.cell(row = 83, column = 7, value = '0,01 - 0,50 %')
	ws.cell(row = 83, column = 8, value = wpd[3][4][6])
	ws.cell(row = 84, column = 7, value = '0,51 - 1,00 %')
	ws.cell(row = 84, column = 8, value = wpd[3][4][7])
	ws.cell(row = 85, column = 7, value = '1,01 - 1,50 %')
	ws.cell(row = 85, column = 8, value = wpd[3][4][8])
	ws.cell(row = 86, column = 7, value = '1,51 - 2,00 %')
	ws.cell(row = 86, column = 8, value = wpd[3][4][9])
	ws.cell(row = 87, column = 7, value = '>2 %')
	ws.cell(row = 87, column = 8, value = wpd[3][4][10])

	return wb

def wpd_sch_sheet(wb, wpd):

	schritte = wpd[2]
	schrittlänge = wpd[3]
	ende = schritte * schrittlänge

	wb.create_sheet('was passiert schleife')
	ws = wb['was passiert schleife']

	ws.cell(row = 1, column = 1, value = 'Was passiert')
	ws.cell(row = 3, column = 1, value = 'Bereich')
	ws.cell(row = 3, column = 2, value = 'Absolute Häufigkeit')
	ws.cell(row = 3, column = 3, value = 'Relative Häufigkeit')
	
	for i in range((wpd[2]+1)*2+1):
		if i == 0:
			name = '<-'+str(ende)
			ws.cell(row = 4+i, column = 1, value = name)
		elif i <= schritte:
			grenze1 = -ende+i*schrittlänge-0.01
			grenze2 = -ende+(i-1)*schrittlänge
			name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
			ws.cell(row = 4+i, column = 1, value = name)
		elif i == schritte+1:
			ws.cell(row = 4+i, column = 1, value = '0')
		elif i <= (schritte+1)*2-1:
			grenze2 = (i-schritte-2)*schrittlänge+0.01
			grenze1 = (i-schritte-1)*schrittlänge
			name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
			ws.cell(row = 4+i, column = 1, value = name)
		else:
			name = '>'+str(ende)
			ws.cell(row = 4+i, column = 1, value = name)

		ws.cell(row = 4+i, column = 2, value = wpd[0][i])
		ws.cell(row = 4+i, column = 3, value = round(wpd[0][i]/sum(wpd[0])*100,2))

	zel = (wpd[2]+1)*2+1 + 1

	for i in range((wpd[2]+1)*2+1):
		if i == 0:
			name = '<-'+str(ende)
			ws.cell(row = 4+zel, column = 1+i*4, value = name)
		elif i <= schritte:
			grenze1 = -ende+i*schrittlänge-0.01
			grenze2 = -ende+(i-1)*schrittlänge
			name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
			ws.cell(row = 4+zel, column = 1+i*4, value = name)
		elif i == schritte+1:
			ws.cell(row = 4+zel, column = 1+i*4, value = '0')
		elif i <= (schritte+1)*2-1:
			grenze2 = (i-schritte-2)*schrittlänge+0.01
			grenze1 = (i-schritte-1)*schrittlänge
			name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
			ws.cell(row = 4+zel, column = 1+i*4, value = name)
		else:
			name = '>'+str(ende)
			ws.cell(row = 4+zel, column = 1+i*4, value = name)

		for j in range((wpd[2]+1)*2+1):

			ws.cell(row = 5+zel, column = 1+i*4, value = 'ein Tag')
			ws.cell(row = 6+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 6+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 6+zel, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 7+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 7+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 7+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 7+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 7+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 7+zel+j, column = 2+i*4, value = wpd[1][0][i][j])
			ws.cell(row = 7+zel+j, column = 3+i*4, value = round(wpd[1][0][i][j]/sum(wpd[1][0][i])*100,2))

		zel2 = zel + (wpd[2]+1)*2+1 + 1 + 2

		for j in range((wpd[2]+1)*2+1):

			ws.cell(row = 5+zel2, column = 1+i*4, value = 'fünf Tag')
			ws.cell(row = 6+zel2, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 6+zel2, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 6+zel2, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 7+zel2+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 7+zel2+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 7+zel2+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 7+zel2+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 7+zel2+j, column = 1+i*4, value = name)

			ws.cell(row = 7+zel2+j, column = 2+i*4, value = wpd[1][1][i][j])
			ws.cell(row = 7+zel2+j, column = 3+i*4, value = round(wpd[1][1][i][j]/sum(wpd[1][1][i])*100,2))

		zel3 = zel2 + (wpd[2]+1)*2+1 + 1 + 2

		for j in range((wpd[2]+1)*2+1):

			ws.cell(row = 5+zel3, column = 1+i*4, value = 'dreißig Tag')
			ws.cell(row = 6+zel3, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 6+zel3, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 6+zel3, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 7+zel3+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 7+zel3+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 7+zel3+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 7+zel3+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 7+zel3+j, column = 1+i*4, value = name)

			ws.cell(row = 7+zel3+j, column = 2+i*4, value = wpd[1][2][i][j])
			ws.cell(row = 7+zel3+j, column = 3+i*4, value = round(wpd[1][2][i][j]/sum(wpd[1][2][i])*100,2))

	return wb

def kurslücken_sheet(wb, ks_pos, ks_neg, verh_ks, nb_ks, b_ks):

	schritte = 5
	schrittlänge = 0.2
	ende = schritte * schrittlänge

	wb.create_sheet('Kurslücken')
	ws = wb['Kurslücken']

	ws.cell(row = 1, column = 1, value = 'Kurslücken')
	ws.cell(row = 3, column = 1, value = 'Positive Kurslücken')
	ws.cell(row = 4, column = 1, value = 'Bestätigte Kurslücken')
	ws.cell(row = 4, column = 5, value = 'Unbestätigte Kurslücken')

	anz_pos = ks_pos[0]
	for j in range(2):
		for i in range(schritte+2):
			if i == 0:
				ws.cell(row = 5+i, column = 1+4*j, value = 'Bereich')
				ws.cell(row = 5+i, column = 2+4*j, value = 'Absolute Häufigkeit')
				ws.cell(row = 5+i, column = 3+4*j, value = 'Relative Häufigkeit')
			elif i <= (schritte):
				grenze2 = (i-1)*schrittlänge+0.01
				grenze1 = (i)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 5+i, column = 1+4*j, value = name)
				ws.cell(row = 5+i, column = 2+4*j, value = anz_pos[j][5-(i-1)])
				ws.cell(row = 5+i, column = 3+4*j, value = anz_pos[j][5-(i-1)]/sum(anz_pos[j]))
			else:
				name = '>'+str(ende)
				ws.cell(row = 5+i, column = 1+4*j, value = name)
				ws.cell(row = 5+i, column = 2+4*j, value = anz_pos[j][5-(i-1)])
				ws.cell(row = 5+i, column = 3+4*j, value = anz_pos[j][5-(i-1)]/sum(anz_pos[j]))

	zel = schritte - 1

	ws.cell(row = 9+zel, column = 1, value = 'Bestätigte Kurslücken')
	
	schritte = 4
	schrittlänge = 0.5
	ende = schritte * schrittlänge
	betit = ['Open-Close','Open-High','Open-Low']
	
	for i in range(3):
		for j in range((schritte+1)*2+1):
			ws.cell(row = 10+zel, column = 1, value = 'Bestätigte Ks-Performance')
			ws.cell(row = 11+zel, column = 1+i*4, value = betit[i])
			ws.cell(row = 12+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 12+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 12+zel, column = 3+i*4, value = 'Relative Häufigkeit')

			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 13+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 13+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 13+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 13+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 13+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 13+zel+j, column = 2+i*4, value = b_ks[i][0][j])
			ws.cell(row = 13+zel+j, column = 3+i*4, value = round(b_ks[i][0][j]/sum(b_ks[i][0])*100,2))


	schritte = 5
	schrittlänge = 0.2
	ende = schritte * schrittlänge
	zel = zel + (schritte+1)*2+3

	for i in range(schritte+1):
		if i <= (schritte-1):
			grenze2 = (i)*schrittlänge+0.01
			grenze1 = (i+1)*schrittlänge
			name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
			ws.cell(row = 10+zel, column = 1+i*4, value = name)
		else:
			ende = (i)*schrittlänge
			name = '>'+str(ende)
			ws.cell(row = 10+zel, column = 1+i*4, value = name)

	schritte = 4
	schrittlänge = 0.5
	ende = schritte * schrittlänge
	perf_pos = ks_pos[1]

	for i in range(6):	
		for j in range((schritte+1)*2+1):

			ws.cell(row = 12+zel, column = 1+i*4, value = 'ein Tag vorher')
			ws.cell(row = 13+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 13+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 13+zel, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 14+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 14+zel+j, column = 2+i*4, value = perf_pos[2][0][5-i][j])
			if sum(perf_pos[2][0][5-i]) > 0:
				ws.cell(row = 14+zel+j, column = 3+i*4, value = round(perf_pos[2][0][5-i][j]/sum(perf_pos[2][0][5-i])*100,2))
			else:
				ws.cell(row = 14+zel+j, column = 3+i*4, value = 'NNN')

	zel = zel + (schritte+1)*2+3

	for i in range(6):	
		for j in range((schritte+1)*2+1):

			ws.cell(row = 12+zel, column = 1+i*4, value = 'gleicher Tag')
			ws.cell(row = 13+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 13+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 13+zel, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 14+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 14+zel+j, column = 2+i*4, value = perf_pos[1][0][5-i][j])
			if sum(perf_pos[1][0][5-i]) > 0:
				ws.cell(row = 14+zel+j, column = 3+i*4, value = round(perf_pos[1][0][5-i][j]/sum(perf_pos[1][0][5-i])*100,2))
			else:
				ws.cell(row = 14+zel+j, column = 3+i*4, value = 'NNN')

	zel = zel + (schritte+1)*2+3

	for i in range(6):	
		for j in range((schritte+1)*2+1):

			ws.cell(row = 12+zel, column = 1+i*4, value = 'ein Tag danach')
			ws.cell(row = 13+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 13+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 13+zel, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 14+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 14+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 14+zel+j, column = 2+i*4, value = perf_pos[0][0][5-i][j])
			if sum(perf_pos[0][0][5-i]) > 0:
				ws.cell(row = 14+zel+j, column = 3+i*4, value = round(perf_pos[0][0][5-i][j]/sum(perf_pos[0][0][5-i])*100,2))
			else:
				ws.cell(row = 14+zel+j, column = 3+i*4, value = 'NNN')

	zel = zel + (schritte+1)*2+1

	ws.cell(row = 15+zel, column = 1, value = 'Teilschluss')
	ws.cell(row = 16+zel, column = 1, value = 'Bereich')
	ws.cell(row = 16+zel, column = 2, value = 'Absolute Häufigkeit')
	ws.cell(row = 16+zel, column = 3, value = 'Relative Häufigkeit')

 
	for j in range(5):
		grenze1 = 0+ j*0.2
		grenze2 = 0.21 + j*0.2
		name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
		ws.cell(row = 17+zel+j, column = 1, value = name)
		ws.cell(row = 17+zel+j, column = 2, value = verh_ks[0][j])
		ws.cell(row = 17 + zel+j, column = 3, value = round(verh_ks[0][j]/sum(verh_ks[0])*100,5))
	
	zel = zel + 9
	schritte = 4
	schrittlänge = 0.5
	ende = schritte * schrittlänge
	betit = ['KsSchluss-Close','Open-Close','Open-High','Open-Low']
	
	for i in range(4):
		for j in range((schritte+1)*2+1):
			ws.cell(row = 14+zel, column = 1, value = 'Unbestätigte Ks-Performance')
			ws.cell(row = 15+zel, column = 1+i*4, value = betit[i])
			ws.cell(row = 16+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 16+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 16+zel, column = 3+i*4, value = 'Relative Häufigkeit')

			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 17+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 17+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 17+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 17+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 17+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 17+zel+j, column = 2+i*4, value = nb_ks[i][0][j])
			ws.cell(row = 17+zel+j, column = 3+i*4, value = round(nb_ks[i][0][j]/sum(nb_ks[i][0])*100,2))


	schritte = 5
	schrittlänge = 0.2
	ende = schritte * schrittlänge
	zel = zel + (schritte+1)*2+2

	ws.cell(row = 16 + zel, column = 1, value = 'Negative Kurslücken')
	ws.cell(row = 17 + zel, column = 1, value = 'Bestätigte Kurslücken')
	ws.cell(row = 17 + zel, column = 5, value = 'Unbestätigte Kurslücken')

	anz_neg = ks_neg[0]

	for j in range(2):
		for i in range(schritte+2):
			if i == 0:
				ws.cell(row = 19+zel+i, column = 1+4*j, value = 'Bereich')
				ws.cell(row = 19+zel+i, column = 2+4*j, value = 'Absolute Häufigkeit')
				ws.cell(row = 19+zel+i, column = 3+4*j, value = 'Relative Häufigkeit')
			elif i <= (schritte):
				grenze2 = -(i-1)*schrittlänge-0.01
				grenze1 = -(i)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 19+zel+i, column = 1+4*j, value = name)
				ws.cell(row = 19+zel+i, column = 2+4*j, value = anz_neg[j][5-(i-1)])
				ws.cell(row = 19+zel+i, column = 3+4*j, value = anz_neg[j][5-(i-1)]/sum(anz_neg[j]))
			else:
				name = '< -'+str(ende)
				ws.cell(row = 19+zel+i, column = 1+4*j, value = name)
				ws.cell(row = 19+zel+i, column = 2+4*j, value = anz_neg[j][5-(i-1)])
				ws.cell(row = 19+zel+i, column = 3+4*j, value = anz_neg[j][5-(i-1)]/sum(anz_neg[j]))

	zel = zel + schritte + 2
	perf_neg = ks_neg[1]
	ws.cell(row = 20+zel, column = 1, value = 'Bestätigte Kurslücken')

	for i in range(schritte+1):
		if i <= (schritte-1):
			grenze2 = -(i)*schrittlänge-0.01
			grenze1 = -(i+1)*schrittlänge
			name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
			ws.cell(row = 21+zel, column = 1+i*4, value = name)
		else:
			ende = (i)*schrittlänge
			name = '< -'+str(ende)
			ws.cell(row = 21+zel, column = 1+i*4, value = name)

	schritte = 4
	schrittlänge = 0.5
	ende = schritte * schrittlänge

	for i in range(6):	
		for j in range((schritte+1)*2+1):

			ws.cell(row = 22+zel, column = 1+i*4, value = 'ein Tag vorher')
			ws.cell(row = 23+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 23+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 23+zel, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 24+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 24+zel+j, column = 2+i*4, value = perf_neg[2][0][5-i][j])
			if sum(perf_pos[2][0][5-i]) > 0:
				ws.cell(row = 24+zel+j, column = 3+i*4, value = round(perf_neg[2][0][5-i][j]/sum(perf_neg[2][0][5-i])*100,2))
			else:
				ws.cell(row = 24+zel+j, column = 3+i*4, value = 'NNN')

	zel = zel + (schritte+1)*2+3

	for i in range(6):	
		for j in range((schritte+1)*2+1):

			ws.cell(row = 22+zel, column = 1+i*4, value = 'gleicher Tag')
			ws.cell(row = 23+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 23+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 23+zel, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 24+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 24+zel+j, column = 2+i*4, value = perf_neg[1][0][5-i][j])
			if sum(perf_pos[1][0][5-i]) > 0:
				ws.cell(row = 24+zel+j, column = 3+i*4, value = round(perf_neg[1][0][5-i][j]/sum(perf_neg[1][0][5-i])*100,2))
			else:
				ws.cell(row = 24+zel+j, column = 3+i*4, value = 'NNN')

	zel = zel + (schritte+1)*2+3


	for i in range(6):	
		for j in range((schritte+1)*2+1):

			ws.cell(row = 22+zel, column = 1+i*4, value = 'ein Tag')
			ws.cell(row = 23+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 23+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 23+zel, column = 3+i*4, value = 'Relative Häufigkeit')
		
			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 24+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 24+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 24+zel+j, column = 2+i*4, value = perf_neg[0][0][5-i][j])
			ws.cell(row = 24+zel+j, column = 3+i*4, value = round(perf_neg[0][0][5-i][j]/sum(perf_neg[0][0][5-i])*100,2))

	zel = zel + (schritte+1)*2+1

	ws.cell(row = 25+zel, column = 1, value = 'Teilschluss')
	ws.cell(row = 26+zel, column = 1, value = 'Bereich')
	ws.cell(row = 26+zel, column = 2, value = 'Absolute Häufigkeit')
	ws.cell(row = 26+zel, column = 3, value = 'Relative Häufigkeit')

 
	for j in range(5):
		grenze1 = 0+ j*0.2
		grenze2 = 0.21 + j*0.2
		name = str(round(grenze1,5))+' ... '+str(round(grenze2,5))+' %'
		ws.cell(row = 27 + zel+j, column = 1, value = name)
		ws.cell(row = 27 + zel+j, column = 2, value = verh_ks[1][j])
		ws.cell(row = 27 + zel+j, column = 3, value = round(verh_ks[1][j]/sum(verh_ks[1])*100,5))

	zel = zel + 7
	schritte = 4
	schrittlänge = 0.5
	ende = schritte * schrittlänge
	betit = ['KsSchluss-Close','Open-Close','Open-High','Open-Low']
	
	for i in range(4):
		for j in range((schritte+1)*2+1):
			ws.cell(row = 26+zel, column = 1, value = 'Unbestätigte Ks-Performance')
			ws.cell(row = 27+zel, column = 1+i*4, value = betit[i])
			ws.cell(row = 28+zel, column = 1+i*4, value = 'Bereich')
			ws.cell(row = 28+zel, column = 2+i*4, value = 'Absolute Häufigkeit')
			ws.cell(row = 28+zel, column = 3+i*4, value = 'Relative Häufigkeit')

			if j == 0:
				name = '<-'+str(ende)
				ws.cell(row = 29+zel+j, column = 1+i*4, value = name)
			elif j <= schritte:
				grenze1 = -ende+j*schrittlänge-0.01
				grenze2 = -ende+(j-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 29+zel+j, column = 1+i*4, value = name)
			elif j == schritte+1:
				ws.cell(row = 29+zel+j, column = 1+i*4, value = '0')
			elif j <= (schritte+1)*2-1:
				grenze2 = (j-schritte-2)*schrittlänge+0.01
				grenze1 = (j-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 29+zel+j, column = 1+i*4, value = name)
			else:
				name = '>'+str(ende)
				ws.cell(row = 29+zel+j, column = 1+i*4, value = name)

			ws.cell(row = 29+zel+j, column = 2+i*4, value = nb_ks[i][1][j])
			ws.cell(row = 29+zel+j, column = 3+i*4, value = round(nb_ks[i][1][j]/sum(nb_ks[i][1])*100,2))

	return wb

def wa_sheet(wb, wa):
	wb.create_sheet('Wochentage')
	ws = wb['Wochentage']

	durch = wa[1]

	ws.cell(row = 1, column = 1, value = 'Wochentage')
	ws.cell(row = 3, column = 1, value = 'Montag')
	ws.cell(row = 3, column = 5, value = 'Dienstag')
	ws.cell(row = 3, column = 9, value = 'Mittwoch')
	ws.cell(row = 3, column = 13, value = 'Donnerstag')
	ws.cell(row = 3, column = 17, value = 'Freitag')

	for i in range(5):
		ws.cell(row = 5, column = 1+4*i, value = 'Performance:')
		ws.cell(row = 6, column = 1+4*i, value = 'Minimal')
		ws.cell(row = 7, column = 1+4*i, value = 'Maximal')
		ws.cell(row = 8, column = 1+4*i, value = 'Durchschnitt')
		ws.cell(row = 6, column = 2+4*i, value = durch[i][0])
		ws.cell(row = 7, column = 2+4*i, value = durch[i][2])
		ws.cell(row = 8, column = 2+4*i, value = round(durch[i][1],5))

	woft = wa[2]
	
	for i in range(5):
		ws.cell(row = 11, column = 1+4*i, value = 'Positiv')
		ws.cell(row = 12, column = 1+4*i, value = 'Neutral')
		ws.cell(row = 13, column = 1+4*i, value = 'Negativ')
		ws.cell(row = 10, column = 2+4*i, value = 'Absolute Häufigkeit')
		ws.cell(row = 11, column = 2+4*i, value = woft[i][2])
		ws.cell(row = 12, column = 2+4*i, value = woft[i][1])
		ws.cell(row = 13, column = 2+4*i, value = woft[i][0])
		ws.cell(row = 10, column = 3+4*i, value = 'Relative Häufigkeit')
		ws.cell(row = 11, column = 3+4*i, value = round(woft[i][2]/sum(woft[i])*100,2))
		ws.cell(row = 12, column = 3+4*i, value = round(woft[i][1]/sum(woft[i])*100,2))
		ws.cell(row = 13, column = 3+4*i, value = round(woft[i][0]/sum(woft[i])*100,2))

	durch_sp = wa[3]
	
	for i in range(5):
		ws.cell(row = 15, column = 1+4*i, value = 'Handelsspanne:')
		ws.cell(row = 16, column = 1+4*i, value = 'Minimal')
		ws.cell(row = 17, column = 1+4*i, value = 'Maximal')
		ws.cell(row = 18, column = 1+4*i, value = 'Durchschnitt')
		ws.cell(row = 16, column = 2+4*i, value = durch_sp[i][0])
		ws.cell(row = 17, column = 2+4*i, value = durch_sp[i][2])
		ws.cell(row = 18, column = 2+4*i, value = round(durch_sp[i][1],2))
	
	wpb = wa[0]
	schritte = 4
	schrittlänge = 0.5
	ende = schritte * schrittlänge

	for j in range(5):
		ws.cell(row = 20, column = 1+4*j, value = 'Performancebereiche:')
		for i in range((4+1)*2+1):
			if i == 0:
				name = '<-'+str(ende)
				ws.cell(row = 21+i, column = 1+4*j, value = name)
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])
			elif i <= schritte:
				grenze1 = -ende+i*schrittlänge-0.01
				grenze2 = -ende+(i-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 21+i, column = 1+4*j, value = name)
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])
			elif i == schritte+1:
				ws.cell(row = 21+i, column = 1+4*j, value = '0')
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])
			elif i <= (schritte+1)*2-1:
				grenze2 = (i-schritte-2)*schrittlänge+0.01
				grenze1 = (i-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 21+i, column = 1+4*j, value = name)
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])
			else:
				name = '>'+str(ende)
				ws.cell(row = 21+i, column = 1+4*j, value = name)
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])

	return wb

def ma_sheet(wb, ma):
	wb.create_sheet('Monate')
	ws = wb['Monate']

	durch = ma[1]

	ws.cell(row = 1, column = 1, value = 'Wochentage')
	ws.cell(row = 3, column = 1, value = 'Januar')
	ws.cell(row = 3, column = 5, value = 'Februar')
	ws.cell(row = 3, column = 9, value = 'März')
	ws.cell(row = 3, column = 13, value = 'April')
	ws.cell(row = 3, column = 17, value = 'Mai')
	ws.cell(row = 3, column = 21, value = 'Juni')
	ws.cell(row = 3, column = 25, value = 'Juli')
	ws.cell(row = 3, column = 29, value = 'August')
	ws.cell(row = 3, column = 33, value = 'September')
	ws.cell(row = 3, column = 37, value = 'Oktober')
	ws.cell(row = 3, column = 41, value = 'November')
	ws.cell(row = 3, column = 45, value = 'Dezember')

	for i in range(12):
		ws.cell(row = 5, column = 1+4*i, value = 'Performance:')
		ws.cell(row = 6, column = 1+4*i, value = 'Minimal')
		ws.cell(row = 7, column = 1+4*i, value = 'Maximal')
		ws.cell(row = 8, column = 1+4*i, value = 'Durchschnitt')
		ws.cell(row = 6, column = 2+4*i, value = durch[i][0])
		ws.cell(row = 7, column = 2+4*i, value = durch[i][2])
		ws.cell(row = 8, column = 2+4*i, value = round(durch[i][1],5))

	woft = ma[2]
	
	for i in range(12):
		ws.cell(row = 11, column = 1+4*i, value = 'Positiv')
		ws.cell(row = 12, column = 1+4*i, value = 'Neutral')
		ws.cell(row = 13, column = 1+4*i, value = 'Negativ')
		ws.cell(row = 10, column = 2+4*i, value = 'Absolute Häufigkeit')
		ws.cell(row = 11, column = 2+4*i, value = woft[i][2])
		ws.cell(row = 12, column = 2+4*i, value = woft[i][1])
		ws.cell(row = 13, column = 2+4*i, value = woft[i][0])
		ws.cell(row = 10, column = 3+4*i, value = 'Relative Häufigkeit')
		ws.cell(row = 11, column = 3+4*i, value = round(woft[i][2]/sum(woft[i])*100,2))
		ws.cell(row = 12, column = 3+4*i, value = round(woft[i][1]/sum(woft[i])*100,2))
		ws.cell(row = 13, column = 3+4*i, value = round(woft[i][0]/sum(woft[i])*100,2))
	
	durch_sp = ma[3]

	for i in range(12):
		ws.cell(row = 15, column = 1+4*i, value = 'Handelsspanne:')
		ws.cell(row = 16, column = 1+4*i, value = 'Minimal')
		ws.cell(row = 17, column = 1+4*i, value = 'Maximal')
		ws.cell(row = 18, column = 1+4*i, value = 'Durchschnitt')
		ws.cell(row = 16, column = 2+4*i, value = durch_sp[i][0])
		ws.cell(row = 17, column = 2+4*i, value = durch_sp[i][2])
		ws.cell(row = 18, column = 2+4*i, value = round(durch_sp[i][1],2))

	wpb = ma[0]
	schritte = 4
	schrittlänge = 0.5
	ende = schritte * schrittlänge

	for j in range(12):
		ws.cell(row = 20, column = 1+4*j, value = 'Performancebereiche:')
		for i in range((4+1)*2+1):
			if i == 0:
				name = '<-'+str(ende)
				ws.cell(row = 21+i, column = 1+4*j, value = name)
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])
			elif i <= schritte:
				grenze1 = -ende+i*schrittlänge-0.01
				grenze2 = -ende+(i-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 21+i, column = 1+4*j, value = name)
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])
			elif i == schritte+1:
				ws.cell(row = 21+i, column = 1+4*j, value = '0')
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])
			elif i <= (schritte+1)*2-1:
				grenze2 = (i-schritte-2)*schrittlänge+0.01
				grenze1 = (i-schritte-1)*schrittlänge
				name = str(round(grenze2,5))+' ... '+str(round(grenze1,5))+' %'
				ws.cell(row = 21+i, column = 1+4*j, value = name)
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])
			else:
				name = '>'+str(ende)
				ws.cell(row = 21+i, column = 1+4*j, value = name)
				ws.cell(row = 21+i, column = 2+4*j, value = wpb[j][i])

	return wb


def workbook_speichern(w, basiswert):
	wb = w
	pfad1 = 'Output'
	pfad2 = basiswert + '_Ebook_Auswertung'
	verzeichnis = os.path.join(pfad1,pfad2)
	
	if not os.path.exists(verzeichnis):
		os.makedirs(verzeichnis)

	wb.save(verzeichnis +'\ ' + basiswert + '_Ebook_Auswertung.xlsx')